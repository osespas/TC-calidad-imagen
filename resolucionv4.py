
# -*- coding: utf-8 -*-
"""
Created on Tue Nov 21 13:22:08 2023
Esta version incorpora tkinter
@author: oestpas
"""

import os, sys, argparse, copy
import pydicom
from pydicom.data import get_testdata_files
import pylibjpeg
import numpy as np
import math
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import cv2
from skimage import measure, io, feature, draw, exposure, color, filters, segmentation
from skimage.filters.rank import maximum
from skimage.measure._regionprops import _RegionProperties
from scipy import ndimage as ndi
from scipy import interpolate, stats
from scipy.fft import fft, fftfreq, ifft, rfft, fftshift, fft2
from scipy.optimize import curve_fit,root
from scipy.interpolate import interp1d
import matplotlib.pylab as pylab
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles.colors import Color
import tkinter as tk
from tkinter import filedialog

def analyze(props):
    
    #Props es un array que arroja informacion de todos los contornos conexos e inconexos que encuentra Canny.
    #La componente-0 es la más grande (círculo interno tal y como está aplicado). PD. se podría hacer igualmente si pillamos el radio externo...
    coords = [props['centroid'][0],props['centroid'][1]] #(row-col)
    radius_row = (props['bbox'][2]-props['bbox'][0])/2
    radius_col = (props['bbox'][3]-props['bbox'][1])/2
    radius = (radius_row + radius_col)/2
    
    return coords, radius

def compute_mtf(x,y,thresolds, debug = False):

    dict_mtf = {}
    for val in thresolds:
        print(' INFO **compute_mtf()** THRESOLDH MTF: ', val)
        xpos = [int(i)for i in x]
        #xpos = list(map(int, x))
        ypos = y
        #print(xpos,ypos)
        y_i = min([i for i in ypos if i >= val])
        x_i = xpos[list(ypos).index(y_i)]
        #print(y_i,x_i)
        y_iplus1 = max([i for i in ypos if i < val])
        x_iplus1 = xpos[list(ypos).index(y_iplus1)]
        #print(y_iplus1,x_iplus1)
        m = (y_iplus1-y_i)/(x_iplus1-x_i)
        if y_i == val:
            mtf = x_i
        else:
            mtf = -(y_i-val)/m + x_i
        print(f" INFO **compute_mtf()** MTF at {val} is ", mtf)
        dict_mtf[val] = mtf
    return dict_mtf 


def create_xlsx(contrast_percentage, dcm_files):
    
    imagenes = dcm_files
    num_imagenes = len(imagenes)
    imagen = {}
    wb = Workbook()
    ws = wb.active
    #ws.append(["Materiales", "\mu (1/cm)", "UH", "Diferencia", "Estado"])

    for k in range(num_imagenes):

        imagen[k] = pydicom.dcmread(imagenes[k], force = True)
        contraste = list(contrast_percentage)
        contraste.insert(0, "Contraste (%)")
        reference = contraste # modificar en un futuro
        state = []
        state.insert(0, "Estado")
        lp = []
        lp.insert(0, " lp/cm")
        fill_color = []
        for i, (val, ref) in enumerate(zip(contraste[1:],reference[1:])):
            result = "Correcto" if ((val-ref)/ref)*100 <= 10 else "Incorrecto"    
            # Set fill color based on the result
            if result == "Correcto":
                fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
            else:
                fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
            fill_color = fill_color + [fill]
            state = state + [result] 
            lp = lp + [i]
        
        ws.append(lp)
        ws.append(contraste)
        ws.append(state)

           
            
    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", 
                               showFirstColumn=False,
                               showLastColumn=False, 
                               showRowStripes=False, 
                               showColumnStripes=False)
    ft = Font(bold=True)
    align = Alignment(horizontal="center", vertical="center", wrapText=True)
    for row in ws["A1:J1"]:
        for cell in row:
            cell.font = ft
    for row in ws["A1:J10"]:
        for cell in row:
            cell.alignment = align
    for row in ws["B3:J3"]:
        for i, cell in enumerate(row):
            cell.fill = fill_color[i]
            
    tab = Table(displayName="Table1", ref="A1:H1")
    
    # set the width of the column 
    for column_cells in ws.columns:
        new_column_length = max(len(str(cell.value)) for cell in column_cells)
        new_column_letter = (get_column_letter(column_cells[0].column))
        if new_column_length > 0:
            ws.column_dimensions[new_column_letter].width = new_column_length*1.23
                      
    tab.tableStyleInfo = style
    ws.add_table(tab)
    
    wb.save("table_resolucion.xlsx")


def Resolucion(head_params, dcm_files):
   
    imagenes = dcm_files
    num_imagenes = len(imagenes)
    mtf_vals = {}
    for k in range(num_imagenes):
        
        # 2. Leer cada imagen       
        imagen = pydicom.dcmread(imagenes[k], force = True)
        #print(imagen)

        pixel_data = imagen.pixel_array
        pixel_data_copy = imagen.pixel_array   
        
        print(' INFO **Resolucion()** LEYENDO IMAGEN: ', imagenes[k])
        
         
        Hospital = imagen.InstitutionName
        ParteDelCuerpo = imagen.BodyPartExamined
        Equipo = imagen.ManufacturerModelName
        TamanoCorte = imagen.SliceThickness
        KV = imagen.KVP
        mAs = imagen.Exposure
        Kernel = imagen.ConvolutionKernel
        TamPixel = imagen.PixelSpacing
        tamPx = TamPixel[0]  # tamaño de Pixel en la exis.
        tamPy = TamPixel[1]  # tamaño de Pixel en la y.
        NumPixelsRows = imagen.Rows
        NumPixelsCols = imagen.Columns
        Intercept = imagen.RescaleIntercept
        Slope = imagen.RescaleSlope
        #pixel_data_copy = Slope*pixel_data_copy+Intercept
        #print(pixel_data_copy.max(), pixel_data.max())
        
        # Get image dimensions
        height, width = pixel_data.shape[:2]	
    	# Compute the center coordinates
        aa = height/2	
        bb = width/2        
        roi_center_settings = {
            "Center": {
                "x": bb,
                "y": aa,
            },
        }
        
        #if args.debug == True:
        #    print(" DEBUG **Resolucion()** Hospital:", Hospital)
        #    print(" DEBUG **Resolucion()** Parte del cuerpo:", ParteDelCuerpo)
        #    print(" DEBUG **Resolucion()** Equipo:", Equipo)
        #    print(" DEBUG **Resolucion()** Tamaño de corte:", TamanoCorte)
        #    print(" DEBUG **Resolucion()** KV:", KV)
        #    print(" DEBUG **Resolucion()** mAs:", mAs)
        #    print(" DEBUG **Resolucion()** Kernel:", Kernel)
            #print(" DEBUG **Resolucion()** Tiempo de Revolución:", TiempoRev)
        #    print(" DEBUG **Resolucion()** Tamaño de Pixel en la x:", tamPx)
        #    print(" DEBUG **Resolucion()** Tamaño de Pixel en la y:", tamPy)
        #    print(" DEBUG **Resolucion()** Número de Pixels por fila:", NumPixelsRows)
        #   print(" DEBUG **Resolucion()** Número de Pixels por columna:", NumPixelsCols)
        #    print(" DEBUG **Resolucion()** HU = a*pixel + b; slope = :", Slope )
        #    print(" DEBUG **Resolucion()** HU = a*pixel + b; intercept = ", Intercept)
      
        # Aplicar el operador Canny para detección de bordes. Scharr es un buen filtro, pero se deja sin detectar uno de los materiales!
        #CABEZA HR40
        if head_params == True:
            sigma = 3
            low_threshold = 5
            high_threshold = 10
            #edges = feature.canny(pixel_data, sigma=3, low_threshold=5, high_threshold=10, use_quantiles = False)
        #TORAX BR40
        else:
            sigma = 3.25
            low_threshold = 5.
            high_threshold = 10
            #edges = feature.canny(pixel_data, sigma=4, low_threshold=7, high_threshold=12, use_quantiles = False)
      
        edges = feature.canny(pixel_data, sigma, low_threshold, high_threshold, use_quantiles = False)
        
        #Radio del perfil
        r = int(47.5/tamPx)

        # Get image dimensions
        height, width = pixel_data.shape[:2]	
    	# Compute the center coordinates
        aa = height/2	
        bb = width/2
        
        fig, ax = plt.subplots()
        plt.imshow(pixel_data, cmap = 'gray') #Show the film at the top
        circle = plt.Circle((bb,aa), r, color = 'red', fill = False)
        ax.add_patch(circle)

        #angles = np.linspace(0, 180, np.round(2*3.1416*r).astype(int))
        angles = np.linspace(-math.pi/2, 3*math.pi/2,360)
        x = aa + r * np.cos(angles)
        y = bb + r * np.sin(angles)
        
      
        # Convert to integer coordinates
        x = np.round(x).astype(int)
        y = np.round(y).astype(int)
        
        profile = pixel_data[x,y]
        #print(profile[359])
        
        #normalized_values = ((profile) / (profile.max())) * 100
        #normalized_values = ((profile - profile.min()) / (profile.max() - profile.min())) * 100
        #De momento no normalizo
 
        
        #Posicion en el perfil de cada grupo de lineas (1 par de lineas, 2 pares, etc...)
        lines_settings = {
             "0": {
                "upper": 15,
                "lower": 0,
             },
             "1": {
                 "upper": 345,
                 "lower": 325,
             },
             "2": {
                 "upper": 317,
                 "lower": 302,
             },
             "3": {
                 "upper": 295,
                 "lower": 278,
             },
             "4": {
                 "upper": 272,
                 "lower": 260,
             },
             "5": {
                 "upper": 252,
                 "lower": 242,
             },
             "6": {
                 "upper": 235,
                 "lower": 224,
             },
             "7": {
                 "upper": 216,
                 "lower": 208,
             },           
             "8": {
                 "upper": 200,
                 "lower": 192,
             },
         }
         
        fig, ax = plt.subplots()
      
        plt.xlabel('Pixels')
        plt.ylabel('Intensidad de pixel')
        plt.grid(linewidth = 1)
        ##plt.xticks(np.arange(0,max(profile),10), rotation = 'vertical')
        #plt.xlim(224,235)
        plt.plot(profile)

        for group in lines_settings.keys():            
            plt.axvline(x =  lines_settings[str(group)]["lower"], color = 'r', linestyle = 'dashed', linewidth = 1)
            plt.axvline(x =  lines_settings[str(group)]["upper"], color = 'r', linestyle = 'dashed', linewidth = 1)
        params = {'axes.labelsize': 18,
          'axes.titlesize': 24,
          'xtick.labelsize':8,
          'ytick.labelsize':16,
          'legend.fontsize': 18}
        plt.rcParams.update(params)

        plt.savefig('resolucionespacial.png')
        plt.show()
        
        #me guardo el rango de cada grupo
        group_profiles = {}
        group_contrast = {}
        for group in lines_settings.keys():
            #print(group)
            if group == "0":
                group_profiles[group] = profile[  lines_settings[group]["lower"]:lines_settings[group]["upper"]   ]
                #print(group_profiles[group])
                group_max = max( group_profiles[group]) #para normalizar al 100 luego
                #print(group_max)
                group_min = np.mean(np.array(group_profiles[group]))
                #print(group_min)
                group_contrast[group] = (group_max - group_min)/(group_max + group_min)
                #print(group_contrast[group])
            else:
                group_profiles[group] = profile[  lines_settings[group]["lower"]:lines_settings[group]["upper"]   ]
                #print(group_profiles[group])
                group_max = max( group_profiles[group])
                group_min = min( group_profiles[group])
                group_contrast[group] = (group_max - group_min)/(group_max + group_min)
                #print(group_contrast[group])
                

        group_max = max( group_profiles["1"]) #para normalizar al 100 luego
        #print(group_max)
        group_min = np.mean(np.array(group_profiles["0"]))
        #print(group_min)
        group_contrast["0"] = (group_max - group_min)/(group_max + group_min)
        #print(group_contrast[group])
                
        
        normalized_values = {}
        mtf = {}
        #https://www.normankoren.com/Tutorials/MTF5.html#quantitative
        #MTF(f) =  ( pi/4 ) * (  100%*C(f)/C(0)  )
        for group in lines_settings.keys():
            normalized_values[group] = np.round(  (group_contrast[group] / group_contrast["0"]) * 100  , 2)
            mtf[group] = np.round( (np.pi/4)*( group_contrast[group] / group_contrast["0"]) * 100  , 2)
       
        print(normalized_values)
        fig, ax = plt.subplots()   
        plt.xlabel('lp/cm')
        plt.ylabel('Contraste (%)')
        plt.ylim(0,100*1.1)
        plt.grid(linewidth = 1)
        
        for group in group_contrast.keys():
            plt.scatter(normalized_values.keys(), normalized_values.values(), color = 'red', marker = 'x')

        plt.savefig('resolucionespacial_contraste.png')
        plt.show()
        
        mtf_thresholds = [50]
        mtf_vals[k] = compute_mtf(list(normalized_values.keys()), list(mtf.values()), mtf_thresholds)

        return list(normalized_values.values())
        

    

def procesar_imagenes():
    path = ruta_texto.get()
    anatomia = seleccion_anatomia.get()
    head_params = anatomia == 'Head'
    
    
    os.chdir(path)
    dcm_files = []
    
    
    for file_name in os.listdir(path):
        try:
            if file_name.endswith(".DCM"):
                dcm_files.append(os.path.join(path, file_name))
                print(' INFO **MAIN()** CARGANDO IMAGEN: ', file_name)
            elif file_name.endswith(".dcm"):
                dcm_files.append(os.path.join(path, file_name))
                print(' INFO **MAIN()** CARGANDO IMAGEN: ', file_name)
        except:
            pass
    
    
    contraste = Resolucion(head_params, dcm_files)
    create_xlsx(contraste, dcm_files)

    # Por ahora, solo imprimiremos los valores seleccionados
    area_mensajes.insert(tk.END, f"Procesando imágenes en {path} para {anatomia}\n")

def seleccionar_carpeta():
    path = filedialog.askdirectory()
    ruta_texto.set(path)



app = tk.Tk()
app.title("Procesador de Imágenes DICOM")

# Ruta de los archivos DICOM
tk.Label(app, text="Ruta de los Archivos DICOM:").pack()
ruta_texto = tk.StringVar()
ruta_entrada = tk.Entry(app, textvariable=ruta_texto, width=50)
ruta_entrada.pack()
tk.Button(app, text="Seleccionar Carpeta", command=seleccionar_carpeta).pack()

# Selección de Anatomía
seleccion_anatomia = tk.StringVar(value='Head')
tk.Radiobutton(app, text="Head", variable=seleccion_anatomia, value='Head').pack()
tk.Radiobutton(app, text="Torax", variable=seleccion_anatomia, value='Torax').pack()

# Botón para procesar imágenes
tk.Button(app, text="Procesar Imágenes", command=procesar_imagenes).pack()

# Área de Mensajes
area_mensajes = tk.Text(app, height=10, width=50)
area_mensajes.pack()

app.mainloop()
        
