# -*- coding: utf-8 -*-
"""
Created on Thu Dec 21 18:15:06 2023

@author: oestpas
"""

# -*- coding: utf-8 -*-
"""
Created on Tue Nov 21 13:22:08 2023

@author: oestpas
"""

import os, sys, argparse, copy
import pydicom
from pydicom.data import get_testdata_files
import pylibjpeg
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import cv2
from skimage import measure, io, feature, draw, exposure, color, filters, segmentation
from skimage.filters.rank import maximum
from skimage.measure._regionprops import _RegionProperties
from scipy import ndimage as ndi
from scipy import interpolate, stats
from scipy.fft import fft, fftfreq, ifft, rfft, fftshift, fft2
from scipy.optimize import curve_fit,root
from scipy.interpolate import interp1d
import matplotlib.pylab as pylab
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles.colors import Color
import tkinter as tk
from tkinter import filedialog

def analyze(props):
    
    #Props es un array que arroja informacion de todos los contornos conexos e inconexos que encuentra Canny.
    #La componente-0 es la más grande (círculo interno tal y como está aplicado). PD. se podría hacer igualmente si pillamos el radio externo...
    coords = [props['centroid'][0],props['centroid'][1]] #(row-col)
    radius_row = (props['bbox'][2]-props['bbox'][0])/2
    radius_col = (props['bbox'][3]-props['bbox'][1])/2
    radius = (radius_row + radius_col)/2
    
    return coords, radius

def create_xlsx(head_params, contrast_percentage, supra_dict,  dcm_files):
    
    imagenes = dcm_files
    num_imagenes = len(imagenes)
    imagen = {}
    wb = Workbook()
    ws = wb.active

    #ws.append(["Materiales", "\mu (1/cm)", "UH", "Diferencia", "Estado"])

    for k in range(num_imagenes):

        imagen[k] = pydicom.dcmread(imagenes[k], force = True)
        contraste = list(contrast_percentage)
        contraste.insert(0, "Contraste (%)")
        #reference = contraste #modificar en un futuro
        sphere = []
        sphere.insert(0, " ID Esfera")
        sphere_sizes = list(supra_dict.keys())
        
        
        sphere_sizes.insert(0, "Tamaño esfera (mm)")
        sphere_contrast = list(supra_dict.values())
        sphere_contrast.insert(0, "Contraste (%)")
        supra_reference = []

        if head_params == "Head":
            supra_reference = ["Supra-Reference", 0.024, 0.021, 0.03, 0.029, 0.023, 0.029, 0.032] #resultados del ERI
        else:
            supra_reference = ["Supra-Reference", 0.082, 0.125, 0.128, 0.057] #resultados del ERI
        
        
        fill_color = []
        state = []
        state.insert(0, "Estado")
        '''
        for i, (val, ref) in enumerate(zip(contraste[1:],reference[1:])):
            result = "Correcto" if ((val-ref)/ref)*100 <= 10 else "Incorrecto"    
            # Set fill color based on the result
            if result == "Correcto":
                fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
            else:
                fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
            #fill_color = fill_color + [fill]
            #state = state + [result] 
            #sphere = sphere + [i]
        '''  
        print(sphere_contrast, supra_reference)
        for i, (val, ref) in enumerate(zip(sphere_contrast[1:],supra_reference[1:])):
            result = "Correcto" if ((val-ref)/ref)*100 <= 10 else "Incorrecto"    
            # Set fill color based on the result
            if result == "Correcto":
                fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
            else:
                fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
            fill_color = fill_color + [fill]
            state = state + [result] 
        
        ws.append(sphere_sizes)
        #ws.append(contraste)
        ws.append(sphere_contrast)
        ws.append(state)

           
            
    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", 
                               showFirstColumn=False,
                               showLastColumn=False, 
                               showRowStripes=False, 
                               showColumnStripes=False)
    ft = Font(bold=True)
    align = Alignment(horizontal="center", vertical="center", wrapText=True)
    for row in ws["A1:N1"]:
        for cell in row:
            cell.font = ft
    for row in ws["A1:N10"]:
        for cell in row:
            cell.alignment = align
    #print(f"B3:H{len(sphere_contrast)-1}")

    for row in ws["B3:E3"]:
        for i, cell in enumerate(row):
            cell.fill = fill_color[i]
            
    tab = Table(displayName="Table1", ref="A1:H1")
    
    # set the width of the column 
    for column_cells in ws.columns:
        new_column_length = max(len(str(cell.value)) for cell in column_cells)
        new_column_letter = (get_column_letter(column_cells[0].column))
        if new_column_length > 0:
            ws.column_dimensions[new_column_letter].width = new_column_length*1.23
                      
    tab.tableStyleInfo = style
    ws.add_table(tab)
    
    wb.save("table_bajocontraste.xlsx")


def BajoContraste(head_params, dcm_files):
   
    imagenes = dcm_files
    num_imagenes = len(imagenes)

    for k in range(num_imagenes):
        
        # 2. Leer cada imagen       
        imagen = pydicom.dcmread(imagenes[k], force = True)

        pixel_data = imagen.pixel_array
        pixel_data_copy = imagen.pixel_array   
        
        print(' INFO **Contraste()** LEYENDO IMAGEN: ', imagenes[k])
        
         
        Hospital = imagen.InstitutionName
       #ParteDelCuerpo = imagen.BodyPartExamined
        Equipo = imagen.ManufacturerModelName
        TamanoCorte = imagen.SliceThickness
        KV = imagen.KVP
        #mAs = imagen.Exposure
        #Kernel = imagen.ConvolutionKernel
        TamPixel = imagen.PixelSpacing
        tamPx = TamPixel[0]  # tamaño de Pixel en la exis.
        tamPy = TamPixel[1]  # tamaño de Pixel en la y.
        NumPixelsRows = imagen.Rows
        NumPixelsCols = imagen.Columns
        Intercept = imagen.RescaleIntercept
        Slope = imagen.RescaleSlope
        #pixel_data_copy = Slope*pixel_data_copy+Intercept
        #print(pixel_data_copy.max(), pixel_data.max())
        plt.imshow(pixel_data, cmap = 'gray')
        
        # Get image dimensions
        height, width = pixel_data.shape[:2]	
    	# Compute the center coordinates
        aa = height/2	
        bb = width/2        
        roi_center_settings = {
            "Center": {
                "x": bb,
                "y": aa,
            },
        }
        
            
        roi_settings = {
            "15mm": {
                "diameter"  : 15,
                "angle"  : -93,
            },
            "9mm": {
                "diameter"  : 9,
                "angle"  : -112,
            },
            "8mm": {
                "diameter"  : 8,
                "angle"  : -129,
            },
            "7mm": {
                "diameter"  : 7,
                "angle"  : -141,
            },
            "6mm": {
                "diameter"  : 6,
                "angle"  : -156 ,
            },
            "5mm": {
                "diameter"  : 5,
                "angle"  : -168,
            },
            "4mm": {
                "diameter"  : 4,
                "angle"  : -177,
            },
            "3mm": {
                "diameter"  : 3,
                "angle"  : 0,
            },
            "2mm": {
                "diameter"  : 2,
                "angle"  : 0,
            },
        }
      
        # Aplicar el operador Canny para detección de bordes. Scharr es un buen filtro, pero se deja sin detectar uno de los materiales!
        #CABEZA HR40
        if head_params == "Head":
            sigma = 3
            low_threshold = 5
            high_threshold = 10
            #edges = feature.canny(pixel_data, sigma=3, low_threshold=5, high_threshold=10, use_quantiles = False)
        #TORAX BR40
        else:
            sigma = 4.6
            low_threshold = 4.
            high_threshold = 7
            #edges = feature.canny(pixel_data, sigma=4, low_threshold=7, high_threshold=12, use_quantiles = False)

        edges = feature.canny(pixel_data, sigma, low_threshold, high_threshold, use_quantiles = False)
        plt.imshow(edges, cmap = 'gray')
        ##mean y otsu son filtros globales, los otros locales
        thresh = np.mean(edges)
        thresh_otsu = filters.threshold_otsu(edges)
        thresh_sauvola = filters.threshold_sauvola(edges)
        thresh_niblack = filters.threshold_niblack(edges, k=0.05)
        
        ##enumero y me quedo con los objetos de tras un procesado a partir del filtro
        bw = edges > thresh
        bw = segmentation.clear_border(bw, buffer_size=int(max(bw.shape) / 50))
        labeled_arr, num_roi = measure.label(bw, return_num=True)
        #print(num_roi)
        #plt.imshow(bw)
        #labeled_arr_scharr, num_roi_scharr = measure.label(edges_scharr, return_num=True)
        #propiedades de los objetos etiquetados
        #objects = measure.regionprops_table(labeled_arr, bw, properties=['eccentricity','area', 'area_filled', 'area_bbox', 'centroid',
                     #'perimeter', 'bbox'])
        objects = measure.regionprops(labeled_arr)
        #objects_scharr = measure.regionprops(labeled_arr_scharr)
        
        ##descarto aquellos circulos que no me interesan a partir de la eccentricidad y ejes
        circle_objects = [obj for obj in objects if obj['eccentricity']<0.9 and obj['axis_major_length']<110/tamPx and obj['area'] > 5 ]
        #circle_objects_scharr = [obj for obj in objects_scharr if obj['centroid'][0] < 200 and obj['centroid'][1] < 300]
        fig, ax = plt.subplots()
        circle = {}
        roi_value = {}
        angle = {}
        coords = {}
        radius = {}
        distance = {}
        for i in range(len(circle_objects)):
            #print(circle_objects[i]['area'])
            #for each object, print out its centroid and print the object itself
            mask = np.zeros(shape=circle_objects[i]['image'].shape)
            mask[circle_objects[i]['image']]=1
            #mostrar que objetos ha detectado
            #plt.figure(figsize=(2,2))
            #plt.imshow(mask, cmap=plt.cm.gray)
            #plt.show()
                    
            
            #Get the coordinates of the circles and plot a rectangle around
            minr, minc, maxr, maxc = circle_objects[i].bbox
            #Apply a red rectangle enclosing each object of interest
            rect = mpatches.Rectangle((minc, minr), maxc - minc, maxr - minr, 
                           fill=False, edgecolor='red', linewidth=2)
            ax.add_patch(rect)
            #wait = input("Press Enter to continue.")          
            
            ###wait = input("Press Enter to continue.")
            #row-col: coords
            coords[i], radius[i] = analyze(circle_objects[i])
            #print (i, coords, radius)
            a = int(coords[i][0])
            b = int(coords[i][1])
            r = int(2/tamPx)

            circle[i] = plt.Circle((b,a), r, color = 'red', fill = False)
            ax.add_patch(circle[i])
            img = np.zeros(shape=pixel_data.shape)
            rr, cc = draw.disk((a, b), r, shape=pixel_data.shape)
            img[rr, cc] = 1
  
            label_image = measure.label(img)
            props = measure.regionprops_table(label_image, pixel_data,
                          properties=['image_intensity', 'intensity_mean'])
                        
            #In HU units    
            roi_value[i] =props['intensity_mean'][0]*Slope + Intercept
            plt.text(b+10,a+10, str(i), color="blue", fontsize=12)
            
            # Get image dimensions
            height, width = pixel_data.shape[:2]	
        	# Compute the center coordinates
            aa = height/2	
            bb = width/2
        
            angle[i] = np.rad2deg(np.arctan2(-a+aa,b-bb))
            distance[i] = (np.sqrt((aa-a)**2+(bb-b)**2))*tamPx

            
           
        #print(roi_dict, distance)
        circle_center = plt.Circle((bb,aa), r, color = 'red', fill = False)
        ax.add_patch(circle_center)
        img = np.zeros(shape=pixel_data.shape)
        rr, cc = draw.disk((aa, bb), r, shape=pixel_data.shape)
        img[rr, cc] = 1
  
        label_image = measure.label(img)
        props = measure.regionprops_table(label_image, pixel_data,
                      properties=['image_intensity', 'intensity_mean'])
                    
        #In HU units    
        circle_center_value =props['intensity_mean'][0]*Slope + Intercept
        
        contrast_percentage = [
            np.round(100/1000 * (( roi_value[i] - circle_center_value) / circle_center_value),3)
            for i in roi_value
        ]
        
      
        for i in range(len(contrast_percentage)):
            print(f' INFO **Contraste()** {i} ROI Contraste Porcentaje ( % )  = ', contrast_percentage[i])
            
        supra_dict = {}
        for roi in roi_settings.keys():
            for i in range(len(roi_value)):
              
                #if np.abs(2*radius[i]*tamPx - roi_settings[roi]["diameter"]) < 1.1 and angle[i] < -90 and distance[i] > 45:
                if  np.abs(angle[i] - roi_settings[roi]["angle"]) < 2.5 :                  
                    supra_dict[roi] = np.round(100/1000 * (( roi_value[i] - circle_center_value) / circle_center_value),3)
                else:
                    pass
        #print(supra_dict)
        plt.imshow(pixel_data, cmap = 'gray')
        plt.savefig('bajo_contraste.png')
        plt.show()
        
        return contrast_percentage, supra_dict
    



def procesar_imagenes():
    path = ruta_texto.get()
    anatomia = seleccion_anatomia.get()

    head_params = anatomia
    
    os.chdir(path)
    dcm_files = []
    
    
    for file_name in os.listdir(path):
        try:
            if file_name.endswith(".DCM"):
                dcm_files.append(os.path.join(path, file_name))
                print(' INFO **MAIN()** CARGANDO IMAGEN: ', file_name)
            elif file_name.endswith(".dcm"):
                dcm_files.append(os.path.join(path, file_name))
                print(' INFO **MAIN()** CARGANDO IMAGEN: ', file_name)
        except:
            pass
    
    
    contrast, supra_dict = BajoContraste(head_params, dcm_files)
    create_xlsx(head_params, contrast, supra_dict, dcm_files)

    # Por ahora, solo imprimiremos los valores seleccionados
    area_mensajes.insert(tk.END, f"Procesando imágenes en {path} para {anatomia}\n")

def seleccionar_carpeta():
    path = filedialog.askdirectory()
    ruta_texto.set(path)
    
    
app = tk.Tk()
app.title("Procesador de Imágenes DICOM")

# Ruta de los archivos DICOM
tk.Label(app, text="Ruta de los Archivos DICOM:").pack()
ruta_texto = tk.StringVar()
ruta_entrada = tk.Entry(app, textvariable=ruta_texto, width=50)
ruta_entrada.pack()
tk.Button(app, text="Seleccionar Carpeta", command=seleccionar_carpeta).pack()

# Selección de Anatomía
seleccion_anatomia = tk.StringVar(value='Head')
tk.Radiobutton(app, text="Head", variable=seleccion_anatomia, value='Head').pack()
tk.Radiobutton(app, text="Torax", variable=seleccion_anatomia, value='Torax').pack()

# Botón para procesar imágenes
tk.Button(app, text="Procesar Imágenes", command=procesar_imagenes).pack()

# Área de Mensajes
area_mensajes = tk.Text(app, height=10, width=50)
area_mensajes.pack()

app.mainloop()