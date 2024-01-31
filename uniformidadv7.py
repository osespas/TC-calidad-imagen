# -*- coding: utf-8 -*-
"""
Created on Tue Nov 21 13:22:08 2023

@author: oestpas
"""

import os, sys, argparse, copy
import pydicom
from pydicom.data import get_testdata_files
import pylibjpeg
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as patch
import cv2
from skimage import measure, io, feature, draw, exposure, color, filters, segmentation
from skimage.filters.rank import maximum
from skimage.measure._regionprops import _RegionProperties
from scipy import ndimage as ndi
from scipy import interpolate
from scipy.fft import fft, fftfreq, ifft, rfft, fftshift, fft2
from scipy.optimize import curve_fit,root
from scipy.interpolate import interp1d
import matplotlib.pylab as pylab
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles.colors import Color
import tkinter as tk
from tkinter import filedialog



def analyze(edges):
    
    #Encontrar los contornos presentes en la imagen (con imagen previamente ajustada)
    contours = measure.find_contours(edges)
    #print(measure.perimeter(edges, neighborhood=4)/(2*np.pi))
    label_image = measure.label(edges)
    props = measure.regionprops_table(label_image, edges,
                  properties=['area', 'area_filled', 'area_bbox', 'centroid',
                              'perimeter', 'bbox'])
    #print(props)
    #Props es un array que arroja informacion de todos los contornos conexos e inconexos que encuentra Canny.
    #La componente-0 es la más grande (círculo interno tal y como está aplicado). PD. se podría hacer igualmente si pillamos el radio externo...
    coords = [props['centroid-0'][0],props['centroid-1'][0]] #(row-col)
    radius_row = (props['bbox-2'][0]-props['bbox-0'][0])/2
    radius_col = (props['bbox-3'][0]-props['bbox-1'][0])/2
    radius = (radius_row + radius_col)/2
    
    return coords, radius
      

def create_xlsx(roi_value, std_value, std_threshold, diffs, dcm_files):
    
    imagenes = dcm_files
    num_imagenes = len(imagenes)
    imagen = {}
    wb = Workbook()
    ws = wb.active
    ws.append(["", "ROI Top", "ROI Right", "ROI Bottom", "ROI Left", "ROI Center"])

    for k in range(num_imagenes):
        # 2. Leer cada imagen
        imagen[k] = pydicom.dcmread(imagenes[k], force = True)
        row = ["Media (UH)"]
        std_row = ["Desviacion Estandard (UH)"]
        diffs.insert(0,"Desviacion Centro (UH)") 
        state = ["Estado"]
        fill_color = []
        for roi in roi_value:
            row = row + [roi_value[roi]]
            std_row = std_row + [std_value[roi]]
            
        for i, val in enumerate(diffs[1:]):
            result = "Correcto" if val <= std_threshold else "Incorrecto"    
            # Set fill color based on the result
            if result == "Correcto":
                fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
            else:
                fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
            fill_color = fill_color + [fill]
            state = state + [result] 

        ws.append(row)
        ws.append(std_row)
        ws.append(diffs)
        ws.append(state)
        
        
    tab = Table(displayName="Table1", ref="A1:G1")
    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", 
                               showFirstColumn=False,
                               showLastColumn=False, 
                               showRowStripes=False, 
                               showColumnStripes=False)
    ft = Font(bold=True)
    align = Alignment(horizontal="center", vertical="center", wrapText=True)
    for row in ws["A1:G1"]:
        for cell in row:
            cell.font = ft
    for row in ws["A1:G10"]:
        for cell in row:
            cell.alignment = align         
    for row in ws["B5:F5"]:
        for i, cell in enumerate(row):
            cell.fill = fill_color[i]
            
    tab = Table(displayName="Table1", ref="A1:H1")
    
    # set the width of the column 
    for column_cells in ws.columns:
        new_column_length = max(len(str(cell.value)) for cell in column_cells)
        new_column_letter = (get_column_letter(column_cells[0].column))
        if new_column_length > 0:
            ws.column_dimensions[new_column_letter].width = new_column_length*1.23
                        

    tab.tableStyleInfo = style
    ws.add_table(tab)
    
    wb.save("table_uniformity.xlsx")
    
def create_xlsx_ruido(std_value, std_threshold,  dcm_files):
    
    imagenes = dcm_files
    num_imagenes = len(imagenes)
    imagen = {}
    wb = Workbook()
    ws = wb.active
    ws.append(["", "ROI Center"])

    for k in range(num_imagenes):
        # 2. Leer cada imagen
        imagen[k] = pydicom.dcmread(imagenes[k], force = True)
        std_row = ["Desviacion Estandard (UH)"]
        state = ["Estado"]
        fill_color = []
        for roi in std_value:
            std_row = std_row + [std_value[roi]]
            
        for i, val in enumerate(std_row[1:]):
            result = "Correcto" if val <= std_threshold else "Incorrecto"    #5 UH para Cabeza; 20 UH para Torax
            # Set fill color based on the result
            if result == "Correcto":
                fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
            else:
                fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
            fill_color = fill_color + [fill]
            state = state + [result] 

        ws.append(std_row)
        ws.append(state)
                        
    tab = Table(displayName="Table1", ref="A1:C4")
    
        
    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", 
                               showFirstColumn=False,
                               showLastColumn=False, 
                               showRowStripes=True, 
                               showColumnStripes=True)
    
    ft = Font(bold=True)
    align = Alignment(horizontal="center", vertical="center", wrapText=True)
    for row in ws["A1:B1"]:
        for cell in row:
            cell.font = ft
    for row in ws["A1:C4"]:
        for cell in row:
            cell.alignment = align         
    for row in ws["B3:B3"]:
        for i, cell in enumerate(row):
            cell.fill = fill_color[i]
    
    # set the width of the column 
    for column_cells in ws.columns:
        new_column_length = max(len(str(cell.value)) for cell in column_cells)
        new_column_letter = (get_column_letter(column_cells[0].column))
        if new_column_length > 0:
            ws.column_dimensions[new_column_letter].width = new_column_length*1.23
            
            
    tab.tableStyleInfo = style
    ws.add_table(tab)
    
    wb.save("table_ruido.xlsx")

def Uniformity(dcm_files):
   
    imagenes = dcm_files
    num_imagenes = len(imagenes)

    for k in range(num_imagenes):
        # 2. Leer cada imagen
        
        imagen = pydicom.dcmread(imagenes[k], force = True)

        pixel_data = imagen.pixel_array
        pixel_data_copy = imagen.pixel_array
        #pixel_data_copy = (np.maximum(pixel_data_copy,0)/pixel_data_copy.max())*100 # entre 0 y 100
        #pixel_data_copy = np.uint8(pixel_data_copy)
        #plt.imshow(pixel_data_copy, cmap = 'gray')
        #HISTOGRAMAS RESCALADOS
        '''
        print(pixel_data.max())
        pixel_data_rescaled = (np.maximum(pixel_data,0)/pixel_data.max())*100 # entre 0 y 100
        #pixel_data_rescaled = np.uint8(pixel_data_rescaled) # integers pixels
        print(pixel_data_rescaled.max())
        
        fig, ax = plt.subplots(2, 1, figsize=(10, 10)) #Create the figures
        ax[0].imshow(pixel_data, cmap = 'gray') #Show the film at the top
        ax[1].imshow(pixel_data_rescaled, cmap = 'gray')
        
        fig2, ax2 = plt.subplots(2, 1, figsize=(10, 10)) #Create the figures
        ax2[0].hist(pixel_data, bins=100, fc='k', ec='k') 
        ax2[1].hist(pixel_data_rescaled, bins=100, fc='k', ec='k') 
        '''        
        
        #Ajustar contraste
        pixel_data_rescaled = exposure.rescale_intensity(pixel_data,  in_range=(np.percentile(pixel_data, 58), np.percentile(pixel_data, 60)))  
        # Aplicar el operador Canny para detección de bordes, tal que el circulo externo desaparezca
        edges = feature.canny(pixel_data_rescaled, sigma=10.0, low_threshold=1500, high_threshold=2000, use_quantiles = False)
        #plt.imshow(edges, cmap = 'gray')
        
        coords, r = analyze(edges)
        a = int(coords[0])
        b = int(coords[1])
        r = int(r)
        
        #GRAFICO
        fig, ax = plt.subplots()
        plt.imshow(pixel_data, cmap = 'gray') #Show the film at the top
        plt.xlabel('')
        plt.ylabel('')
        #plt.scatter(a,b, c='black', marker='x', s=60) 
        circ = plt.Circle((a,b), r, color = 'black', fill = False)
        ax.add_patch(circ)
        #plt.show()


        print(' INFO **Uniformity()** LEYENDO IMAGEN: ', imagenes[k])

        
        Hospital = imagen.InstitutionName
        ParteDelCuerpo = imagen.BodyPartExamined
        Equipo = imagen.ManufacturerModelName
        TamanoCorte = imagen.SliceThickness
        KV = imagen.KVP
        mAs = imagen.Exposure
        Kernel = imagen.ConvolutionKernel
        TamPixel = imagen.PixelSpacing
        tamPx = TamPixel[0]  # tamaño de Pixel en la exis.
        tamPy = TamPixel[1]  # tamaño de Pixel en la y.
        NumPixelsRows = imagen.Rows
        NumPixelsCols = imagen.Columns
        Intercept = imagen.RescaleIntercept
        Slope = imagen.RescaleSlope
        #pixel_data_copy = Slope*pixel_data_copy+Intercept
        #print(pixel_data_copy.max(), pixel_data.max())
        
   
        #Protocolo Español: radio de 1 cm
        radius = int(10/tamPx) #en pixeles
        radiusy = int(10/tamPy) #en pixeles
        
        #Ojo con la notacion: ROI's a 1 cm del borde
        roi_settings = {
            "Top": {
                "x": b,
                "y": a+r-2*radius,
            },
            "Right": {
                "x": b+r-2*radius,
                "y": a,
            },
            "Bottom": {
                "x": b,
                "y": a-r+2*radius,

            },
            "Left": {
                "x": b-r+2*radius,
                "y": a,

            },
            "Center": {
                "x": b,
                "y": a,
            },
        }
           
        roi_value = {}
        circle_plt = {}
        for roi in roi_settings.keys():
           
            # Calculate the coordinates of the circle ROI        
            roi_x = roi_settings[roi]['x'] 
            roi_y = roi_settings[roi]['y'] 
                    
  
            circle_plt[roi] = plt.Circle((roi_y,roi_x), radius, color = 'black', fill = False)
            ax.add_patch(circle_plt[roi])
            
        #plt.show()
        #plt.imshow(pixel_data_copy, cmap = 'gray')
        plt.savefig('uniformidad.png')
        #plt.imshow(pixel_data, cmap = 'gray')

        
        circle = {}
        std_value = {}
        #fig, ax = plt.subplots(figsize=(10, 6))
        for roi in roi_settings.keys():
            # Calculate the coordinates of the circle ROI        
            roi_x = roi_settings[roi]['x'] 
            roi_y = roi_settings[roi]['y']
            #roi_x = Slope*roi_x+Intercept
            #roi_y = Slope*roi_y+Intercept
            #mask = np.zeros_like(pixel_data_copy, dtype=np.uint8)
            
            #circle[roi] = cv2.circle(pixel_data_copy, (roi_y,roi_x), radius,  (0,255,0), 1)
            #circle[roi] = segmentation.clear_border(circle[roi])
            #cv2.addWeighted(pixel_data_copy,1, circle[roi],0,0)
            
            circle[roi] = plt.Circle((roi_y,roi_x), radius, color = 'red', fill = False)
            ax.add_patch(circle[roi])
            img = np.zeros(shape=pixel_data.shape)
            rr, cc = draw.disk((roi_x, roi_y), radius, shape=pixel_data.shape)
            img[rr, cc] = 1
            
            label_image = measure.label(img)
            #label_image = measure.label(circle[roi])
            ##props = measure.regionprops_table(label_image, circle[roi],
            ##              properties=['intensity_mean'])
            props = measure.regionprops_table(label_image, pixel_data_copy,
                          properties=['intensity_mean'])
            
            std = pixel_data_copy[rr,cc]
            #print(std)
            #print(props['intensity_mean'])
            
            #plt.show()
            #plt.imshow(label_image, cmap = 'gray')
            #plt.savefig('%s.png' %(imagenes[k]))
            #In HU units     
            #roi_value[roi] = np.mean(circle[roi])
            #roi_value[roi] = (np.mean(circle[roi])*Slope) + Intercept
            roi_value[roi] = np.round(props['intensity_mean'][0]*Slope + Intercept,2)
            std_value[roi] = np.round(np.std(std),2)
            #print(np.mean(circle[roi]))
            print(f' DEBUG **Uniformity()** {roi} ROI Value: ', roi_value[roi])
            print(f' DEBUG **Uniformity()** {roi} ROI Std. Value: ', std_value[roi])
           

        
        #std_value = np.std(roi_value)
        #if args.debug== True: print(f' DEBUG **Uniformity()** ROI Std. Value: ', std_value)
            
        uis = [
            100 * (1- np.abs((roi_value[roi] - roi_value['Center']) / (roi_value['Center'] + 1000)))
            for roi in roi_settings.keys()
        ]
       
        abs_uis = np.abs(uis)
        ui = uis[np.argmax(abs_uis)]
        print(' INFO **Uniformity()** Uniformity Index= ', ui)
        
        #Checkear que |ROI(x) - ROI(Center)| < 5 HU
        diffs = [
            (roi_value[roi] - roi_value['Center'])
            for roi in roi_settings.keys()
        ]
        
        abs_diffs = np.abs(diffs)
        diff = abs_diffs[np.argmax(abs_diffs)]
        print(f' INFO **Uniformity()** Maximum Difference ROI Value wrt ROI Center: ', diff)
        if diff < 5:
            print(f' INFO **Uniformity()** IN TOLERANCE ')
        else:
            print(f' INFO **Uniformity()** OUT OF TOLERANCE ')
            
        return roi_value, std_value, diffs
    

def Ruido(std_threshold, dcm_files):
   
    imagenes = dcm_files
    num_imagenes = len(imagenes)

    for k in range(num_imagenes):
        # 2. Leer cada imagen
        
        imagen = pydicom.dcmread(imagenes[k], force = True)

        pixel_data = imagen.pixel_array
        pixel_data_copy = imagen.pixel_array
 
        #Ajustar contraste
        pixel_data_rescaled = exposure.rescale_intensity(pixel_data,  in_range=(np.percentile(pixel_data, 58), np.percentile(pixel_data, 60)))  
        # Aplicar el operador Canny para detección de bordes, tal que el circulo externo desaparezca
        edges = feature.canny(pixel_data_rescaled, sigma=10.0, low_threshold=1500, high_threshold=2000, use_quantiles = False)
        #plt.imshow(edges, cmap = 'gray')
        
        coords, r = analyze(edges)
        a = int(coords[0])
        b = int(coords[1])
        r = int(r)
        
        #GRAFICO
        fig, ax = plt.subplots()
        plt.imshow(pixel_data, cmap = 'gray') #Show the film at the top
        plt.xlabel('')
        plt.ylabel('')
        #plt.scatter(a,b, c='black', marker='x', s=60) 
        circ = plt.Circle((a,b), r, color = 'black', fill = False)
        ax.add_patch(circ)
        #plt.show()


        print(' INFO **Ruido()** LEYENDO IMAGEN: ', imagenes[k])

        
        Hospital = imagen.InstitutionName
        ParteDelCuerpo = imagen.BodyPartExamined
        Equipo = imagen.ManufacturerModelName
        TamanoCorte = imagen.SliceThickness
        KV = imagen.KVP
        mAs = imagen.Exposure
        Kernel = imagen.ConvolutionKernel
        TamPixel = imagen.PixelSpacing
        tamPx = TamPixel[0]  # tamaño de Pixel en la exis.
        tamPy = TamPixel[1]  # tamaño de Pixel en la y.
        NumPixelsRows = imagen.Rows
        NumPixelsCols = imagen.Columns
        Intercept = imagen.RescaleIntercept
        Slope = imagen.RescaleSlope
        #pixel_data_copy = Slope*pixel_data_copy+Intercept
        #print(pixel_data_copy.max(), pixel_data.max())

        
        #Protocolo Español: radio de 1 cm
        radius = int(np.sqrt((500/tamPx**2)/np.pi)) #en pixeles
        radiusy = int(np.sqrt((500/tamPy**2)/np.pi)) #en pixeles
        
        #Ojo con la notacion: ROI's a 1 cm del borde
        roi_settings = {
            "Center": {
                "x": b,
                "y": a,
            },
        }
           
        roi_value = {}
        circle_plt = {}
        for roi in roi_settings.keys():
           
            # Calculate the coordinates of the circle ROI        
            roi_x = roi_settings[roi]['x'] 
            roi_y = roi_settings[roi]['y'] 
                    
  
            circle_plt[roi] = plt.Circle((roi_y,roi_x), radius, color = 'black', fill = False)
            ax.add_patch(circle_plt[roi])
            
        #plt.show()
        #plt.imshow(pixel_data_copy, cmap = 'gray')
        plt.savefig('ruido.png')
        #plt.imshow(pixel_data, cmap = 'gray')

        
        circle = {}
        std_value = {}
        #fig, ax = plt.subplots(figsize=(10, 6))
        for roi in roi_settings.keys():
            # Calculate the coordinates of the circle ROI        
            roi_x = roi_settings[roi]['x'] 
            roi_y = roi_settings[roi]['y']
            #roi_x = Slope*roi_x+Intercept
            #roi_y = Slope*roi_y+Intercept
            #mask = np.zeros_like(pixel_data_copy, dtype=np.uint8)
            
            #circle[roi] = cv2.circle(pixel_data_copy, (roi_y,roi_x), radius,  (0,255,0), 1)
            #circle[roi] = segmentation.clear_border(circle[roi])
            #cv2.addWeighted(pixel_data_copy,1, circle[roi],0,0)
            
            circle[roi] = plt.Circle((roi_y,roi_x), radius, color = 'red', fill = False)
            ax.add_patch(circle[roi])
            img = np.zeros(shape=pixel_data.shape)
            rr, cc = draw.disk((roi_x, roi_y), radius, shape=pixel_data.shape)
            img[rr, cc] = 1
            
            label_image = measure.label(img)
            #label_image = measure.label(circle[roi])
            ##props = measure.regionprops_table(label_image, circle[roi],
            ##              properties=['intensity_mean'])
            props = measure.regionprops_table(label_image, pixel_data_copy,
                          properties=['intensity_mean'])
            
            std = pixel_data_copy[rr,cc]
            #print(std)
            #print(props['intensity_mean'])
            
            #plt.show()
            #plt.imshow(label_image, cmap = 'gray')
            #plt.savefig('%s.png' %(imagenes[k]))
            #In HU units     
            #roi_value[roi] = np.mean(circle[roi])
            #roi_value[roi] = (np.mean(circle[roi])*Slope) + Intercept
            roi_value[roi] = np.round(props['intensity_mean'][0]*Slope + Intercept,2)
            std_value[roi] = np.round(np.std(std),2)
            #print(np.mean(circle[roi]))
            print(f' DEBUG **Ruido()** {roi} ROI Value: ', roi_value[roi])
            print(f' DEBUG **Ruido()** {roi} ROI Std. Value: ', std_value[roi])
                        
            if std_value[roi] < std_threshold:
                print(f' INFO **Ruido()** IN TOLERANCE ')
            else:
                print(f' INFO **Ruido()** OUT OF TOLERANCE ')
           

        
            
        return roi_value, std_value
            
        
def procesar_imagenes():
    path = ruta_texto.get()
    anatomia = seleccion_anatomia.get()

    head_params = anatomia
    
    os.chdir(path)
    dcm_files = []
    
    
    for file_name in os.listdir(path):
        try:
            if file_name.endswith(".DCM"):
                dcm_files.append(os.path.join(path, file_name))
                print(' INFO **MAIN()** CARGANDO IMAGEN: ', file_name)
            elif file_name.endswith(".dcm"):
                dcm_files.append(os.path.join(path, file_name))
                print(' INFO **MAIN()** CARGANDO IMAGEN: ', file_name)
        except:
            pass
    

    std_threshold = 0
    if head_params == 'Head': std_threshold = 5
    elif head_params == 'Torax': std_threshold = 20 
    
    roi_value, std_value, diffs = Uniformity(dcm_files)
    roi_value_ruido, std_value_ruido= Ruido(std_threshold, dcm_files)
    create_xlsx(roi_value, std_value, std_threshold, diffs, dcm_files)
    create_xlsx_ruido(std_value_ruido, std_threshold, dcm_files)

    # Por ahora, solo imprimiremos los valores seleccionados
    area_mensajes.insert(tk.END, f"Procesando imágenes en {path} para {anatomia}\n")

def seleccionar_carpeta():
    path = filedialog.askdirectory()
    ruta_texto.set(path)
    
    
app = tk.Tk()
app.title("Procesador de Imágenes DICOM")

# Ruta de los archivos DICOM
tk.Label(app, text="Ruta de los Archivos DICOM:").pack()
ruta_texto = tk.StringVar()
ruta_entrada = tk.Entry(app, textvariable=ruta_texto, width=50)
ruta_entrada.pack()
tk.Button(app, text="Seleccionar Carpeta", command=seleccionar_carpeta).pack()

# Selección de Anatomía
seleccion_anatomia = tk.StringVar(value='Head')
tk.Radiobutton(app, text="Head", variable=seleccion_anatomia, value='Head').pack()
tk.Radiobutton(app, text="Torax", variable=seleccion_anatomia, value='Torax').pack()

# Botón para procesar imágenes
tk.Button(app, text="Procesar Imágenes", command=procesar_imagenes).pack()

# Área de Mensajes
area_mensajes = tk.Text(app, height=10, width=50)
area_mensajes.pack()

app.mainloop()


     