# -*- coding: utf-8 -*-
"""
Created on Thu Dec 21 18:15:06 2023

@author: oestpas
"""

# -*- coding: utf-8 -*-
"""
Created on Tue Nov 21 13:22:08 2023

@author: madiamarf
"""

import os, sys, argparse, copy
import pydicom
from pydicom.data import get_testdata_files
import pylibjpeg
import numpy as np
import math
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
#import cv2
from skimage import measure, io, feature, draw, exposure, color, filters, segmentation
from skimage.filters.rank import maximum
from skimage.measure._regionprops import _RegionProperties
from scipy import ndimage as ndi
from scipy import interpolate, stats
from scipy.fft import fft, fftfreq, ifft, rfft, fftshift, fft2
from scipy.optimize import curve_fit,root
from scipy.interpolate import interp1d
import matplotlib.pylab as pylab
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles.colors import Color
import tkinter as tk
from tkinter import filedialog
from tkinter import ttk
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import inspect
from datetime import datetime


#######################################
#
#            Funciones para Resolución
#
#######################################
def analyze(props):
    
    #Props es un array que arroja informacion de todos los contornos conexos e inconexos que encuentra Canny.
    #La componente-0 es la más grande (círculo interno tal y como está aplicado). PD. se podría hacer igualmente si pillamos el radio externo...
    coords = [props['centroid'][0],props['centroid'][1]] #(row-col)
    radius_row = (props['bbox'][2]-props['bbox'][0])/2
    radius_col = (props['bbox'][3]-props['bbox'][1])/2
    radius = (radius_row + radius_col)/2
    
    return coords, radius

def compute_mtf(x,y,thresolds, debug = False):

    dict_mtf = {}
    for val in thresolds:
        print(' INFO **compute_mtf()** THRESOLDH MTF: ', val)
        xpos = [int(i)for i in x]
        #xpos = list(map(int, x))
        ypos = y
        #print(xpos,ypos)
        y_i = min([i for i in ypos if i >= val])
        x_i = xpos[list(ypos).index(y_i)]
        #print(y_i,x_i)
        y_iplus1 = max([i for i in ypos if i < val])
        x_iplus1 = xpos[list(ypos).index(y_iplus1)]
        #print(y_iplus1,x_iplus1)
        m = (y_iplus1-y_i)/(x_iplus1-x_i)
        if y_i == val:
            mtf = x_i
        else:
            mtf = -(y_i-val)/m + x_i
        print(f" INFO **compute_mtf()** MTF at {val} is ", mtf)
        dict_mtf[val] = mtf
    return dict_mtf 


def create_xlsx(contrast_percentage, dcm_files):
    
    imagenes = dcm_files
    num_imagenes = len(imagenes)
    imagen = {}
    wb = Workbook()
    ws = wb.active
    #ws.append(["Materiales", "\mu (1/cm)", "UH", "Diferencia", "Estado"])

    for k in range(num_imagenes):

        imagen[k] = pydicom.dcmread(imagenes[k], force = True)
        contraste = list(contrast_percentage)
        contraste.insert(0, "Contraste (%)")
        reference = contraste # modificar en un futuro
        state = []
        state.insert(0, "Estado")
        lp = []
        lp.insert(0, " lp/cm")
        fill_color = []
        for i, (val, ref) in enumerate(zip(contraste[1:],reference[1:])):
            result = "Correcto" if ((val-ref)/ref)*100 <= 10 else "Incorrecto"    
            # Set fill color based on the result
            if result == "Correcto":
                fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
            else:
                fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
            fill_color = fill_color + [fill]
            state = state + [result] 
            lp = lp + [i]
        
        ws.append(lp)
        ws.append(contraste)
        ws.append(state)

           
            
    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", 
                               showFirstColumn=False,
                               showLastColumn=False, 
                               showRowStripes=False, 
                               showColumnStripes=False)
    ft = Font(bold=True)
    align = Alignment(horizontal="center", vertical="center", wrapText=True)
    for row in ws["A1:J1"]:
        for cell in row:
            cell.font = ft
    for row in ws["A1:J10"]:
        for cell in row:
            cell.alignment = align
    for row in ws["B3:J3"]:
        for i, cell in enumerate(row):
            cell.fill = fill_color[i]
            
    tab = Table(displayName="Table1", ref="A1:H1")
    
    # set the width of the column 
    for column_cells in ws.columns:
        new_column_length = max(len(str(cell.value)) for cell in column_cells)
        new_column_letter = (get_column_letter(column_cells[0].column))
        if new_column_length > 0:
            ws.column_dimensions[new_column_letter].width = new_column_length*1.23
                      
    tab.tableStyleInfo = style
    ws.add_table(tab)
    
    wb.save("table_resolucion.xlsx")
    
    
        #6-2-2024
def update_treeview(treeview, data_to_export):
    
    global last_function_called
    
    # Limpia el Treeview
    for i in treeview.get_children():
        treeview.delete(i)
    
    # Datos de ejemplo, reemplaza con los valores reales
    cols = range(len(data_to_export))
    #estados = ["Correcto" if val >= 10 else "Incorrecto" for val in data_to_export]
    
    # Inserta los datos en el Treeview en funcion de la ultima funcion llamada
    for i in cols:
        if last_function_called == procesar_resolucion:
            treeview.insert('', 'end', values=(i, data_to_export[i]))
        elif last_function_called == procesar_contraste:
            materials = ["Air", "PMP", "LDPE", "Poly", "Acrylic", "Delrin", "Teflon", "Fecha"]
            treeview.insert('', 'end', values=(materials[i], data_to_export[i]))
        #6-2-2024


def Resolucion(head_params, dcm_files):
   
    imagenes = dcm_files
    num_imagenes = len(imagenes)
    mtf_vals = {}
    for k in range(num_imagenes):
        
        # 2. Leer cada imagen       
        imagen = pydicom.dcmread(imagenes[k], force = True)
        #print(imagen)

        pixel_data = imagen.pixel_array
        pixel_data_copy = imagen.pixel_array   
        
        print(' INFO **Resolucion()** LEYENDO IMAGEN: ', imagenes[k])
        
         
        Hospital = imagen.InstitutionName
        ParteDelCuerpo = imagen.BodyPartExamined
        Equipo = imagen.ManufacturerModelName
        TamanoCorte = imagen.SliceThickness
        KV = imagen.KVP
        mAs = imagen.Exposure
        Kernel = imagen.ConvolutionKernel
        TamPixel = imagen.PixelSpacing
        tamPx = TamPixel[0]  # tamaño de Pixel en la exis.
        tamPy = TamPixel[1]  # tamaño de Pixel en la y.
        NumPixelsRows = imagen.Rows
        NumPixelsCols = imagen.Columns
        Intercept = imagen.RescaleIntercept
        Slope = imagen.RescaleSlope
        #pixel_data_copy = Slope*pixel_data_copy+Intercept
        #print(pixel_data_copy.max(), pixel_data.max())
        
        # Get image dimensions
        height, width = pixel_data.shape[:2]	
    	# Compute the center coordinates
        aa = height/2	
        bb = width/2        
        roi_center_settings = {
            "Center": {
                "x": bb,
                "y": aa,
            },
        }
        
        #if args.debug == True:
        #    print(" DEBUG **Resolucion()** Hospital:", Hospital)
        #    print(" DEBUG **Resolucion()** Parte del cuerpo:", ParteDelCuerpo)
        #    print(" DEBUG **Resolucion()** Equipo:", Equipo)
        #    print(" DEBUG **Resolucion()** Tamaño de corte:", TamanoCorte)
        #    print(" DEBUG **Resolucion()** KV:", KV)
        #    print(" DEBUG **Resolucion()** mAs:", mAs)
        #    print(" DEBUG **Resolucion()** Kernel:", Kernel)
            #print(" DEBUG **Resolucion()** Tiempo de Revolución:", TiempoRev)
        #    print(" DEBUG **Resolucion()** Tamaño de Pixel en la x:", tamPx)
        #    print(" DEBUG **Resolucion()** Tamaño de Pixel en la y:", tamPy)
        #    print(" DEBUG **Resolucion()** Número de Pixels por fila:", NumPixelsRows)
        #   print(" DEBUG **Resolucion()** Número de Pixels por columna:", NumPixelsCols)
        #    print(" DEBUG **Resolucion()** HU = a*pixel + b; slope = :", Slope )
        #    print(" DEBUG **Resolucion()** HU = a*pixel + b; intercept = ", Intercept)
      
        # Aplicar el operador Canny para detección de bordes. Scharr es un buen filtro, pero se deja sin detectar uno de los materiales!
        #CABEZA HR40
        if head_params == True:
            sigma = 3
            low_threshold = 5
            high_threshold = 10
            #edges = feature.canny(pixel_data, sigma=3, low_threshold=5, high_threshold=10, use_quantiles = False)
        #TORAX BR40
        else:
            sigma = 3.25
            low_threshold = 5.
            high_threshold = 10
            #edges = feature.canny(pixel_data, sigma=4, low_threshold=7, high_threshold=12, use_quantiles = False)
      
        edges = feature.canny(pixel_data, sigma, low_threshold, high_threshold, use_quantiles = False)
        
        #Radio del perfil
        r = int(47.5/tamPx)

        # Get image dimensions
        height, width = pixel_data.shape[:2]	
    	# Compute the center coordinates
        aa = height/2	
        bb = width/2
        
        fig, ax = plt.subplots()
        plt.imshow(pixel_data, cmap = 'gray') #Show the film at the top
        circle = plt.Circle((bb,aa), r, color = 'red', fill = False)
        ax.add_patch(circle)

        #angles = np.linspace(0, 180, np.round(2*3.1416*r).astype(int))
        angles = np.linspace(-math.pi/2, 3*math.pi/2,360)
        x = aa + r * np.cos(angles)
        y = bb + r * np.sin(angles)
        
      
        # Convert to integer coordinates
        x = np.round(x).astype(int)
        y = np.round(y).astype(int)
        
        profile = pixel_data[x,y]
        #print(profile[359])
        
        #normalized_values = ((profile) / (profile.max())) * 100
        #normalized_values = ((profile - profile.min()) / (profile.max() - profile.min())) * 100
        #De momento no normalizo
 
        
        #Posicion en el perfil de cada grupo de lineas (1 par de lineas, 2 pares, etc...)
        lines_settings = {
             "0": {
                "upper": 15,
                "lower": 0,
             },
             "1": {
                 "upper": 345,
                 "lower": 325,
             },
             "2": {
                 "upper": 317,
                 "lower": 302,
             },
             "3": {
                 "upper": 295,
                 "lower": 278,
             },
             "4": {
                 "upper": 272,
                 "lower": 260,
             },
             "5": {
                 "upper": 252,
                 "lower": 242,
             },
             "6": {
                 "upper": 235,
                 "lower": 224,
             },
             "7": {
                 "upper": 216,
                 "lower": 208,
             },           
             "8": {
                 "upper": 200,
                 "lower": 192,
             },
         }
         
        fig, ax = plt.subplots()
      
        plt.xlabel('Pixels')
        plt.ylabel('Intensidad de pixel')
        plt.grid(linewidth = 1)
        ##plt.xticks(np.arange(0,max(profile),10), rotation = 'vertical')
        #plt.xlim(224,235)
        plt.plot(profile)

        for group in lines_settings.keys():            
            plt.axvline(x =  lines_settings[str(group)]["lower"], color = 'r', linestyle = 'dashed', linewidth = 1)
            plt.axvline(x =  lines_settings[str(group)]["upper"], color = 'r', linestyle = 'dashed', linewidth = 1)
        params = {'axes.labelsize': 18,
          'axes.titlesize': 24,
          'xtick.labelsize':8,
          'ytick.labelsize':16,
          'legend.fontsize': 18}
        plt.rcParams.update(params)

        plt.savefig('resolucionespacial.png')
        plt.show()
        
        #me guardo el rango de cada grupo
        group_profiles = {}
        group_contrast = {}
        for group in lines_settings.keys():
            #print(group)
            if group == "0":
                group_profiles[group] = profile[  lines_settings[group]["lower"]:lines_settings[group]["upper"]   ]
                #print(group_profiles[group])
                group_max = max( group_profiles[group]) #para normalizar al 100 luego
                #print(group_max)
                group_min = np.mean(np.array(group_profiles[group]))
                #print(group_min)
                group_contrast[group] = (group_max - group_min)/(group_max + group_min)
                #print(group_contrast[group])
            else:
                group_profiles[group] = profile[  lines_settings[group]["lower"]:lines_settings[group]["upper"]   ]
                #print(group_profiles[group])
                group_max = max( group_profiles[group])
                group_min = min( group_profiles[group])
                group_contrast[group] = (group_max - group_min)/(group_max + group_min)
                #print(group_contrast[group])
                

        group_max = max( group_profiles["1"]) #para normalizar al 100 luego
        #print(group_max)
        group_min = np.mean(np.array(group_profiles["0"]))
        #print(group_min)
        group_contrast["0"] = (group_max - group_min)/(group_max + group_min)
        #print(group_contrast[group])
                
        
        normalized_values = {}
        mtf = {}
        #https://www.normankoren.com/Tutorials/MTF5.html#quantitative
        #MTF(f) =  ( pi/4 ) * (  100%*C(f)/C(0)  )
        for group in lines_settings.keys():
            normalized_values[group] = np.round(  (group_contrast[group] / group_contrast["0"]) * 100  , 2)
            mtf[group] = np.round( (np.pi/4)*( group_contrast[group] / group_contrast["0"]) * 100  , 2)
       
        print(normalized_values)
        fig, ax = plt.subplots()   
        plt.xlabel('lp/cm')
        plt.ylabel('Contraste (%)')
        plt.ylim(0,100*1.1)
        plt.grid(linewidth = 1)
        
        for group in group_contrast.keys():
            plt.scatter(normalized_values.keys(), normalized_values.values(), color = 'red', marker = 'x')

        plt.savefig('resolucionespacial_contraste.png')
        plt.show()
        
        mtf_thresholds = [50]
        mtf_vals[k] = compute_mtf(list(normalized_values.keys()), list(mtf.values()), mtf_thresholds)

        return list(normalized_values.values())
        
#######################################
#
#            Funciones para Contraste
#
#######################################
def Contraste(head_params, dcm_files):
   
    imagenes = dcm_files
    num_imagenes = len(imagenes)

    for k in range(num_imagenes):
        
        # 2. Leer cada imagen       
        imagen = pydicom.dcmread(imagenes[k], force = True)

        pixel_data = imagen.pixel_array
        pixel_data_copy = imagen.pixel_array   
        
        print(' INFO **Contraste()** LEYENDO IMAGEN: ', imagenes[k])
        
         
        Hospital = imagen.InstitutionName
        ParteDelCuerpo = imagen.BodyPartExamined
        Equipo = imagen.ManufacturerModelName
        TamanoCorte = imagen.SliceThickness
        KV = imagen.KVP
        mAs = imagen.Exposure
        Kernel = imagen.ConvolutionKernel
        TamPixel = imagen.PixelSpacing
        tamPx = TamPixel[0]  # tamaño de Pixel en la exis.
        tamPy = TamPixel[1]  # tamaño de Pixel en la y.
        NumPixelsRows = imagen.Rows
        NumPixelsCols = imagen.Columns
        Intercept = imagen.RescaleIntercept
        Slope = imagen.RescaleSlope
        #pixel_data_copy = Slope*pixel_data_copy+Intercept
        #print(pixel_data_copy.max(), pixel_data.max())
         
        

        AIR = -1000
        PMP = -200
        LDPE = -100
        POLY = -35
        ACRYLIC = 120
        DELRIN = 340
        TEFLON = 990
        WATER = 0
        #Coeficientes de atenuacion lineal en cm-1 tomados a 70 kEV en base al protocolo
        roi_settings = {
            "Air": {
                "value": AIR,
                "angle": -90,
                "mu"   : 0,
                "color": "red",
                "theo" : "-1046:-986",
                "ref"  : -1000,
            },
            "PMP": {
                "value": PMP,
                "angle": -60,
                "mu"   : 0.157,
                "color": "blue",
                "theo" : "-220:-172",
                "ref"  : -200,
            },
            "LDPE": {
                "value": LDPE,
                "angle": 0,
                "mu"   : 0.174,
                "color": "orange",
                "theo" : "-121:-87",
                "ref"  : -100,
            },
            "Poly": {
                "value": POLY,
                "angle": 60,
                "mu"   : 0.188,
                "color": "yellow",
                "theo" : "-65:-29",
                "ref"  : -35,
            },
            "Acrylic": {
                "value": ACRYLIC,
                "angle": 120,
                "mu"   : 0.215,
                "color": "pink",
                "theo" : "92:137",
                "ref"  : 120,
            },
            "Delrin": {
                "value": DELRIN,
                "angle": -180, #tb 180 depende de si da error
                "mu"   : 0.245,
                "color": "black",
                "theo" : "315:345",
                "ref"  : 340,
            },
            "Teflon": {
                "value": TEFLON,
                "angle": -120,
                "mu"   : 0.363,
                "color": "cyan",
                "theo" : "915:955",
                "ref"  : 990,
            },
        }
        
        #CABEZA HR40
        if head_params == "Head":
            sigma = 5
            low_threshold = 1
            high_threshold = 10
        #TORAX BR40
        else:
            sigma = 5
            low_threshold = 1
            high_threshold = 10

        edges = feature.canny(pixel_data, sigma, low_threshold, high_threshold, use_quantiles = False)
        # Aplicar el operador Canny para detección de bordes. Scharr es un buen filtro, pero se deja sin detectar uno de los materiales!
        #edges = feature.canny(pixel_data, sigma=5.0, low_threshold=1, high_threshold=10, use_quantiles = False)

        plt.imshow(edges, cmap = 'gray')
        
        ##mean y otsu son filtros globales, los otros locales
        thresh = np.mean(edges)
        thresh_otsu = filters.threshold_otsu(edges)
        thresh_sauvola = filters.threshold_sauvola(edges)
        thresh_niblack = filters.threshold_niblack(edges, k=0.05)
        
        ##enumero y me quedo con los objetos de tras un procesado a partir del filtro
        bw = edges > thresh
        bw = segmentation.clear_border(bw, buffer_size=int(max(bw.shape) / 50))
        labeled_arr, num_roi = measure.label(bw, return_num=True)
        #labeled_arr_scharr, num_roi_scharr = measure.label(edges_scharr, return_num=True)
        #propiedades de los objetos etiquetados
        #objects = measure.regionprops_table(labeled_arr, bw, properties=['eccentricity','area', 'area_filled', 'area_bbox', 'centroid',
                     #'perimeter', 'bbox'])
        objects = measure.regionprops(labeled_arr)
        #objects_scharr = measure.regionprops(labeled_arr_scharr)
        
        ##descarto aquellos circulos que no me interesan a partir de la eccentricidad y ejes
        circle_objects = [obj for obj in objects if obj['eccentricity']<0.7 and obj['axis_major_length']>12/tamPx and obj['axis_major_length']<110/tamPx]
        #circle_objects_scharr = [obj for obj in objects_scharr if obj['centroid'][0] < 200 and obj['centroid'][1] < 300]
        fig, ax = plt.subplots(figsize=(10, 6))
        circle = {}
        roi_value = {}
        std_value = {}
        angle = {}
        for i in range(len(circle_objects)):

            #for each object, print out its centroid and print the object itself
            mask = np.zeros(shape=circle_objects[i]['image'].shape)
            mask[circle_objects[i]['image']]=1
            #mostrar que objetos ha detectado
            '''
            plt.figure(figsize=(2,2))
            plt.imshow(mask, cmap=plt.cm.gray)
            plt.show()
            '''
            
            #Get the coordinates of the circles and plot a rectangle around
            minr, minc, maxr, maxc = circle_objects[i].bbox
            #Apply a red rectangle enclosing each object of interest
            rect = mpatches.Rectangle((minc, minr), maxc - minc, maxr - minr, 
                           fill=False, edgecolor='red', linewidth=2)
            ax.add_patch(rect)
            ###wait = input("Press Enter to continue.")
            #row-col: coords
            coords, radius = analyze(circle_objects[i])
            ##print (i, coords, radius)
            a = int(coords[0])
            b = int(coords[1])
            r = int(3.5/tamPx)

            circle[i] = plt.Circle((b,a), r, color = "red", fill = False)
            ax.add_patch(circle[i])
            img = np.zeros(shape=pixel_data.shape)
            rr, cc = draw.disk((a, b), r, shape=pixel_data.shape)
            img[rr, cc] = 1
            
            std = pixel_data_copy[rr,cc]
            
            label_image = measure.label(img)
            props = measure.regionprops_table(label_image, pixel_data,
                          properties=['image_intensity', 'intensity_mean'])
                        
            #In HU units    
            roi_value[i] = np.round(props['intensity_mean'][0]*Slope + Intercept,2)
            std_value[i] = np.round(np.std(std),2)
            plt.text(b,a, str(i), color="red", fontsize=12)
            
            
            # Get image dimensions
            height, width = pixel_data.shape[:2]	
        	# Compute the center coordinates
            aa = height/2	
            bb = width/2
            #print(aa,bb)
            #print(a,b)         
            angle[i] = np.rad2deg(np.arctan2(-a+aa,b-bb))


        plt.imshow(pixel_data, cmap = 'gray')    
        
        roi_dict = {}
        std_dict = {}
        for roi in roi_settings.keys():
            for i in range(len(roi_value)):
                if np.abs(-angle[i]+int(roi_settings[roi]['angle'])) < 1:
                    roi_dict[roi] = roi_value[i]
                    std_dict[roi] = std_value[i]

                else:
                    pass

        print(" INFO **Contraste()** Valores ROI = ", roi_dict)
        print(" INFO **Contraste()** Std. Values = ", std_dict)
        
        
        return roi_dict, std_dict, roi_settings

           

def plot_linearity(roi_value, std_value, roi_settings):
    
    fig = Figure(figsize=(12, 8))
    
    ax = fig.add_subplot(111)
    ax.set_ylim(-1500, 1500)
    ax.set_xlim(0, 0.4)
    ax.set_xlabel('Atenuación ($\mu = cm^{-1}$)')
    ax.set_ylabel('UH')
    ax.set_title("Linealidad Contraste")

    t = np.linspace(0, 600, 50)
    mu = {}
    
    for roi in roi_settings.keys():
        ax.scatter(roi_settings[roi]['mu'], roi_value[roi], label=roi, color=roi_settings[roi]["color"], marker='x')
        mu[roi] = roi_settings[roi]['mu']
    
    slope, intercept, r, p, se = stats.linregress(list(mu.values()), list(roi_value.values()))
    ax.plot(t, intercept + slope * t, 'red', label='Ajuste')
    ax.legend()
    
    results = ['Escala de Contraste = %.6f UH/$\mathrm{cm}^{-1}$' % (1/slope)]
    textstr = '\n'.join(results)
    ax.text(0.25, 0.55, textstr, fontsize='large')

    return fig

    
def create_xlsx2(roi_value, roi_settings, dcm_files):
    
    imagenes = dcm_files
    num_imagenes = len(imagenes)
    imagen = {}
    wb = Workbook()
    ws = wb.active
    #ws.append(["Materiales", "\mu (1/cm)", "UH", "Diferencia", "Estado"])

    for k in range(num_imagenes):

        imagen[k] = pydicom.dcmread(imagenes[k], force = True)
        materials = list(roi_value.keys())
        materials.insert(0, "Materiales")
        mu_theo = ["Coef. Atenuacion (1/cm)"]
        hu = list(roi_value.values())
        hu.insert(0, "HU")
        hu_theo = ["HU teorico"]
        hu_ref = ["HU referencia"]
        hu_diff = ["Diferencia HU"]
        state = ["Estado"]
        hu_lower = []
        hu_upper = []
        fill_color = []
        for roi in roi_value.keys():
            mu_theo = mu_theo + [roi_settings[roi]["mu"]]
            hu_theo = hu_theo + [roi_settings[roi]["theo"]]
            hu_ref = hu_ref + [roi_settings[roi]["ref"]]
            hu_lower = hu_lower + [roi_settings[roi]["theo"].split(":")[0]]
            hu_upper = hu_upper + [roi_settings[roi]["theo"].split(":")[1]]
        
        hu_diff = hu_diff + list(map(lambda a, b: np.round(a-b,2), hu[1:], hu_ref[1:]))
        #hu_diff = hu_diff + [a-b for a,b in zip(hu_ref, hu_theo)]

        for i, val in enumerate(hu[1:]):
            result = "Correcto" if int(val) >= int(hu_lower[i]) and int(val) <= int(hu_upper[i]) else "Incorrecto"    
            # Set fill color based on the result
            if result == "Correcto":
                fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
            else:
                fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
            fill_color = fill_color + [fill]
            state = state + [result] 
        
        ws.append(materials)
        ws.append(mu_theo)
        ws.append(hu)
        ws.append(hu_theo)
        ws.append(hu_ref)
        ws.append(hu_diff)
        ws.append(state)
           
            
    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", 
                               showFirstColumn=False,
                               showLastColumn=False, 
                               showRowStripes=False, 
                               showColumnStripes=False)
    ft = Font(bold=True)
    align = Alignment(horizontal="center", vertical="center", wrapText=True)
    for row in ws["A1:H1"]:
        for cell in row:
            cell.font = ft
    for row in ws["A1:H10"]:
        for cell in row:
            cell.alignment = align
    for row in ws["B7:H7"]:
        for i, cell in enumerate(row):
            cell.fill = fill_color[i]
            
    tab = Table(displayName="Table1", ref="A1:H1")
    
    # set the width of the column 
    for column_cells in ws.columns:
        new_column_length = max(len(str(cell.value)) for cell in column_cells)
        new_column_letter = (get_column_letter(column_cells[0].column))
        if new_column_length > 0:
            ws.column_dimensions[new_column_letter].width = new_column_length*1.23
                      
    tab.tableStyleInfo = style
    ws.add_table(tab)
    
    wb.save("table_contraste.xlsx")
    
    return hu[1:]
    
    #######################################
#
#            Funciones para botones
#
#######################################
    

def procesar_resolucion():
    # 01-04-24 exportar en funcion de la funcion que se ha llamado
    global last_function_called, data_to_export
    last_function_called = procesar_resolucion
    # 01-04-24 exportar en funcion de la funcion que se ha llamado
            
    path = ruta_texto.get()
    anatomia = seleccion_anatomia.get()
    head_params = anatomia == 'Head'
    
    
    os.chdir(path)
    dcm_files = []
    
    
    for file_name in os.listdir(path):
        try:
            if file_name.endswith(".DCM"):
                dcm_files.append(os.path.join(path, file_name))
                print(' INFO **MAIN()** CARGANDO IMAGEN: ', file_name)
            elif file_name.endswith(".dcm"):
                dcm_files.append(os.path.join(path, file_name))
                print(' INFO **MAIN()** CARGANDO IMAGEN: ', file_name)
        except:
            pass
    
    
    contraste = Resolucion(head_params, dcm_files)
    create_xlsx(contraste, dcm_files)
    
    # Limpia el área de mensajes y el Treeview si ya contiene datos
    area_mensajes.delete('1.0', tk.END)
    #for i in treeview.get_children():
    #    treeview.delete(i)
    # Si ya existe un canvas, primero lo eliminas
    for widget in grafico_frame.winfo_children():
        widget.destroy()
    
    #6-2-2024
    # Define el Treeview en GUI
    columns = ("lp/cm", "Contraste (%)")
    treeview = ttk.Treeview(app, columns=columns, show='headings')
    for col in columns:
        treeview.heading(col, text=col)
        treeview.column(col, anchor="center")
    treeview.grid(row=7, column=0, columnspan=2, sticky='nsew')
    #6-2-2024
    
    # 01-04-24 exportar a excel
    data_to_export = contraste
    # 01-04-24 exportar a excel

    #6-2-2024
    update_treeview(treeview, contraste)
    #6-2-2024

    # Imprimiremos los valores seleccionados
    area_mensajes.insert(tk.END, f"Procesando imágenes en {path} para {anatomia}\n")


def procesar_contraste():
    # 01-04-24 exportar en funcion de la funcion que se ha llamado
    global last_function_called, data_to_export
    last_function_called = procesar_contraste
    # 01-04-24
    
    
    path = ruta_texto.get()
    anatomia = seleccion_anatomia.get()
    head_params = anatomia == 'Head'
    
    
    os.chdir(path)
    dcm_files = []
    
    
    for file_name in os.listdir(path):
        try:
            if file_name.endswith(".DCM"):
                dcm_files.append(os.path.join(path, file_name))
                print(' INFO **MAIN()** CARGANDO IMAGEN: ', file_name)
            elif file_name.endswith(".dcm"):
                dcm_files.append(os.path.join(path, file_name))
                print(' INFO **MAIN()** CARGANDO IMAGEN: ', file_name)
        except:
                pass
        
    ##########################

    roi_dict, std_dict, roi_settings = Contraste(head_params, dcm_files)
    fig = plot_linearity(roi_dict, std_dict, roi_settings)
    hu = create_xlsx2(roi_dict, roi_settings, dcm_files)
    
    # Limpia el área de mensajes y el Treeview si ya contiene datos
    area_mensajes.delete('1.0', tk.END)
    #for i in treeview_hu.get_children():
    #    treeview_hu.delete(i)

    # Si ya existe un canvas, primero lo eliminas
    for widget in grafico_frame.winfo_children():
        widget.destroy()
    
    canvas = FigureCanvasTkAgg(fig, master=grafico_frame)
    canvas.draw()
    canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)
    canvas.get_tk_widget().config(width=600, height=500)
        
        
    #03-04-2024
    # Define el Treeview en GUI
    columns_hu = ("Materiales", "HU")
    treeview_hu = ttk.Treeview(app, columns=columns_hu, show='headings')
    for col in columns_hu:
        treeview_hu.heading(col, text=col)
        treeview_hu.column(col, anchor="center")
    treeview_hu.grid(row=7, column=0, columnspan=2, sticky='nsew')
    #03-04-2024
    
    # 01-04-24 exportar a excel
    data_to_export = hu
    # 01-04-24 exportar a excel
    update_treeview(treeview_hu, hu)
    # Imprimiremos los valores seleccionados
    area_mensajes.insert(tk.END, f"Procesando imágenes en {path} para {anatomia}\n")




def seleccionar_carpeta():
    path = filedialog.askdirectory()
    ruta_texto.set(path)


def exportar_a_excel():

    path = ruta_texto.get()
    anatomia = seleccion_anatomia.get()
    head_params = anatomia == 'Head'

    # 01-04-24 exportar en funcion de la funcion que se ha llamado
    global last_function_called, data_to_export
    if last_function_called is None:
        print("No function has been called yet.")
        return

    # Prepare data based on the last function called
    if last_function_called == procesar_resolucion:
        print("Data from procesar_resolucion")
        if data_to_export[0] != "Contraste (%)":
            data_to_export.insert(0, "Contraste (%)")
        current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        data_to_export.append(current_date)
        lpcm = ["lp/cm", 0, 1 , 2, 3, 4, 5, 6, 7, 8, "Fecha"]
        print(data_to_export)
        try:
            wb = load_workbook("historico_resolucion.xlsx")
        except FileNotFoundError:
            wb = Workbook()

        ws = wb.active       
        row = 1    
        for i, value in enumerate(lpcm, start=1):
            ws.cell(row=row, column=i, value=value)

        # Encontrar la primera fila vacía en la columna A
        row = 2
        while ws.cell(row=row, column=1).value is not None: 
            row += 1
        for i, value in enumerate(data_to_export, start=1):
            ws.cell(row=row, column=i, value=value)
        wb.save("historico_resolucion.xlsx")
        
    elif last_function_called == procesar_contraste:
        print("Data from procesar_contraste")
        
        if data_to_export[0] != "HU":
            data_to_export.insert(0, "HU")
        current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        data_to_export.append(current_date)
        materials = ["Materiales", "Air", "PMP", "LDPE", "Poly", "Acrylic", "Delrin", "Teflon", "Fecha"]
        print(data_to_export)
        try:
            wb = load_workbook("historico_contraste.xlsx")
        except FileNotFoundError:
            wb = Workbook()

        ws = wb.active       
        row = 1    
        for i, value in enumerate(materials, start=1):
            ws.cell(row=row, column=i, value=value)

        # Encontrar la primera fila vacía en la columna A
        row = 2
        while ws.cell(row=row, column=1).value is not None: 
            row += 1
        for i, value in enumerate(data_to_export, start=1):
            ws.cell(row=row, column=i, value=value)
        wb.save("historico_contraste.xlsx")
        

    else:
        print("Waiting for data")
    # 01-04-24 exportar en funcion de la funcion que se ha llamado

#####################################################################            
######################APP GUI #######################################
#####################################################################

global last_function_called

app = tk.Tk()
app.title("Procesador de Imágenes DICOM")

# Ruta de los archivos DICOM
tk.Label(app, text="Ruta de los Archivos DICOM:").grid(row=0, column=0, columnspan=2)
ruta_texto = tk.StringVar()
ruta_entrada = tk.Entry(app, textvariable=ruta_texto, width=100)
ruta_entrada.grid(row=1, column=0, columnspan=2)
tk.Button(app, text="Seleccionar Carpeta", command=seleccionar_carpeta).grid(row=2, column=0, columnspan=2)

# Selección de Anatomía
seleccion_anatomia = tk.StringVar(value='Head')
tk.Radiobutton(app, text="Head", variable=seleccion_anatomia, value='Head').grid(row=3, column=0)
tk.Radiobutton(app, text="Torax", variable=seleccion_anatomia, value='Torax').grid(row=3, column=1)

# Botón para procesar imágenes para Resolución
tk.Button(app, text="Procesar Resolución", command=procesar_resolucion).grid(row=4, column=0)

# Botón para procesar imágenes para Contraste
tk.Button(app, text="Procesar Contraste", command=procesar_contraste).grid(row=4, column=1)

#Botón para exportar a Excel histórico
tk.Button(app, text="Exportar", command=exportar_a_excel).grid(row=4, column=3)

# Área de Mensajes
area_mensajes = tk.Text(app, height=10, width=100)
area_mensajes.grid(row=5, column=0, columnspan=2)

# Contenedor para el gráfico
grafico_frame = tk.Frame(app)
grafico_frame.grid(row=6, column=0, columnspan=2)


app.mainloop()
        
