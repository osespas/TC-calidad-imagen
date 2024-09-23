 # -*- coding: utf-8 -*-
"""
Created on Thu Dec 21 18:15:06 2023

@author: oestpas
"""

# -*- coding: utf-8 -*-
"""
Created on Tue Nov 21 13:22:08 2023

@author: madiamarf
"""

import os, sys, argparse, copy
import pydicom
from pydicom.data import get_testdata_files
import pylibjpeg
import numpy as np
import math
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
#import cv2
from skimage import measure, io, feature, draw, exposure, color, filters, segmentation
from skimage.filters.rank import maximum
from skimage.measure._regionprops import _RegionProperties
from scipy import ndimage as ndi
from scipy import interpolate, stats
from scipy.fft import fft, fftfreq, ifft, rfft, fftshift, fft2
from scipy.optimize import curve_fit,root
from scipy.interpolate import interp1d
import matplotlib.pylab as pylab
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles.colors import Color
import tkinter as tk
from tkinter import filedialog
from tkinter import ttk
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import inspect
from datetime import datetime


#######################################
#
#            Funciones para Resolución
#
#######################################
def analyze(props):
    
    #Props es un array que arroja informacion de todos los contornos conexos e inconexos que encuentra Canny.
    coords = [props['centroid'][0],props['centroid'][1]] #(row-col)
    radius_row = (props['bbox'][2]-props['bbox'][0])/2
    radius_col = (props['bbox'][3]-props['bbox'][1])/2
    radius = (radius_row + radius_col)/2
    
    return coords, radius

def compute_mtf(x,y,thresolds, debug = False):

    dict_mtf = {}
    for val in thresolds:
        print(' INFO **compute_mtf()** THRESOLDH MTF: ', val)
        xpos = [int(i)for i in x]
        #xpos = list(map(int, x))
        ypos = y
        #print(xpos,ypos)
        y_i = min([i for i in ypos if i >= val])
        x_i = xpos[list(ypos).index(y_i)]
        #print(y_i,x_i)
        y_iplus1 = max([i for i in ypos if i < val])
        x_iplus1 = xpos[list(ypos).index(y_iplus1)]
        #print(y_iplus1,x_iplus1)
        m = (y_iplus1-y_i)/(x_iplus1-x_i)
        if y_i == val:
            mtf = x_i
        else:
            mtf = -(y_i-val)/m + x_i
        print(f" INFO **compute_mtf()** MTF at {val} is ", np.abs(mtf))
        dict_mtf[val] = np.abs(mtf)
    return dict_mtf 

def fit_splines(x,xp,y):
       splines = interpolate.splrep(x, y)
       fxp = interpolate.splev(xp, splines)
       return fxp
def create_xlsx(contrast_percentage, dcm_files):
    
    imagenes = dcm_files
    num_imagenes = len(imagenes)
    imagen = {}
    wb = Workbook()
    ws = wb.active
    #ws.append(["Materiales", "\mu (1/cm)", "UH", "Diferencia", "Estado"])

    for k in range(num_imagenes):

        imagen[k] = pydicom.dcmread(imagenes[k], force = True)
        contraste = list(contrast_percentage)
        contraste.insert(0, "Contraste (%)")
        reference = contraste # modificar en un futuro
        state = []
        state.insert(0, "Estado")
        lp = []
        lp.insert(0, " lp/cm")
        fill_color = []
        for i, (val, ref) in enumerate(zip(contraste[1:],reference[1:])):
            result = "Correcto" if ((val-ref)/ref)*100 <= 10 else "Incorrecto"    
            # Set fill color based on the result
            if result == "Correcto":
                fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
            else:
                fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
            fill_color = fill_color + [fill]
            state = state + [result] 
            lp = lp + [i]
        
        ws.append(lp)
        ws.append(contraste)
        ws.append(state)

           
            
    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", 
                               showFirstColumn=False,
                               showLastColumn=False, 
                               showRowStripes=False, 
                               showColumnStripes=False)
    ft = Font(bold=True)
    align = Alignment(horizontal="center", vertical="center", wrapText=True)
    for row in ws["A1:J1"]:
        for cell in row:
            cell.font = ft
    for row in ws["A1:J10"]:
        for cell in row:
            cell.alignment = align
    for row in ws["B3:J3"]:
        for i, cell in enumerate(row):
            cell.fill = fill_color[i]
            
    tab = Table(displayName="Table1", ref="A1:H1")
    
    # set the width of the column 
    for column_cells in ws.columns:
        new_column_length = max(len(str(cell.value)) for cell in column_cells)
        new_column_letter = (get_column_letter(column_cells[0].column))
        if new_column_length > 0:
            ws.column_dimensions[new_column_letter].width = new_column_length*1.23
                      
    tab.tableStyleInfo = style
    ws.add_table(tab)
    
    wb.save("table_resolucion.xlsx")
    
    
        #6-2-2024
def update_treeview(treeview, keys, data_to_export):
    
    global last_function_called, rois_bajocontraste
    
    # Limpia el Treeview
    for i in treeview.get_children():
        treeview.delete(i)
        
    # Datos de ejemplo, reemplaza con los valores reales
    cols = range(len(data_to_export))
    #estados = ["Correcto" if val >= 10 else "Incorrecto" for val in data_to_export]

    # Inserta los datos en el Treeview en funcion de la ultima funcion llamada
    for i in cols:
        if last_function_called == procesar_resolucion:
            treeview.insert('', 'end', values=(i, data_to_export[i]))
        elif last_function_called == procesar_mtf:
            mtf = keys
            treeview.insert('', 'end', values=(mtf[i]*100, round(data_to_export[i],2)))
        elif last_function_called == procesar_contraste:
            materials = keys
            treeview.insert('', 'end', values=(materials[i], data_to_export[i]))
        elif last_function_called == procesar_uniformidad:
            rois = keys
            treeview.insert('', 'end', values=(rois[i], data_to_export[i]))
        elif last_function_called == procesar_bajocontraste:
            rois_bajocontraste = keys
            treeview.insert('', 'end', values=(rois_bajocontraste[i], data_to_export[i]))
        #6-2-2024
        
    for item in treeview.get_children():
        values = treeview.item(item, "values")
        if not any(values):  # Eliminar si todos los valores están vacíos
            treeview.delete(item)


def Resolucion(head_params, dcm_files):
   
    imagenes = dcm_files
    num_imagenes = len(imagenes)
    mtf_vals = {}
    for k in range(num_imagenes):
        
        # 2. Leer cada imagen       
        imagen = pydicom.dcmread(imagenes[k], force = True)
        #print(imagen)

        pixel_data = imagen.pixel_array
        pixel_data_copy = imagen.pixel_array   
        
        print(' INFO **Resolucion()** LEYENDO IMAGEN: ', imagenes[k])
        
         
        Hospital = imagen.InstitutionName
        ParteDelCuerpo = imagen.BodyPartExamined
        Equipo = imagen.ManufacturerModelName
        TamanoCorte = imagen.SliceThickness
        KV = imagen.KVP
        mAs = imagen.Exposure
        Kernel = imagen.ConvolutionKernel
        TamPixel = imagen.PixelSpacing
        tamPx = TamPixel[0]  # tamaño de Pixel en la exis.
        tamPy = TamPixel[1]  # tamaño de Pixel en la y.
        NumPixelsRows = imagen.Rows
        NumPixelsCols = imagen.Columns
        Intercept = imagen.RescaleIntercept
        Slope = imagen.RescaleSlope
        #pixel_data_copy = Slope*pixel_data_copy+Intercept
        #print(pixel_data_copy.max(), pixel_data.max())
        
        # Get image dimensions
        height, width = pixel_data.shape[:2]	
    	# Compute the center coordinates
        aa = height/2	
        bb = width/2        
        roi_center_settings = {
            "Center": {
                "x": bb,
                "y": aa,
            },
        }
                
        #Radio del perfil
        r = int(47.5/tamPx)

        # Get image dimensions
        height, width = pixel_data.shape[:2]	
    	# Compute the center coordinates
        #OSC: Si el fantoma está un poco girado hay que sumarle y restarle al centro de la imagen para que pase por todas las barras.
        #CT4: aa +3.5; bb -1.5
        #CT3: aa +3.5; bb + 3.5
        if Equipo == "SOMATOM X.cite":
            aa = height/2 
            bb = width/2  
        elif Equipo == "SOMATOM Definition AS+":
            aa = height/2 +3.5
            bb = width/2 + 3.5
        elif Equipo == "Spectral CT":
            aa = height/2 +3.5
            bb = width/2 -1.5
        elif Equipo == "Ingenuity CT":
            aa = height/2 +6
            bb = width/2 
        
        fig, ax = plt.subplots()
        plt.imshow(pixel_data, cmap = 'gray') #Show the film at the top
        circle = plt.Circle((bb,aa), r, color = 'red', fill = False)
        ax.add_patch(circle)

        #angles = np.linspace(0, 180, np.round(2*3.1416*r).astype(int))
        angles = np.linspace(-math.pi/2, 3*math.pi/2,360)
        x = aa + r * np.cos(angles)
        y = bb + r * np.sin(angles)
        
      
        # Convert to integer coordinates
        x = np.round(x).astype(int)
        y = np.round(y).astype(int)
        
        profile = pixel_data[x,y]
 
        
        #Posicion en el perfil de cada grupo de lineas (1 par de lineas, 2 pares, etc...)
        lines_settings = {
             "0": {
                "upper": 15,
                "lower": 0,
             },
             "1": {
                 "upper": 345,
                 "lower": 325,
             },
             "2": {
                 "upper": 317,
                 "lower": 302,
             },
             "3": {
                 "upper": 295,
                 "lower": 278,
             },
             "4": {
                 "upper": 272,
                 "lower": 260,
             },
             "5": {
                 "upper": 252,
                 "lower": 242,
             },
             "6": {
                 "upper": 235,
                 "lower": 224,
             },
             "7": {
                 "upper": 216,
                 "lower": 208,
             },           
             "8": {
                 "upper": 200,
                 "lower": 192,
             },
         }
         
        fig, ax = plt.subplots()
      
        plt.xlabel('Pixels')
        plt.ylabel('Intensidad de pixel')
        plt.grid(linewidth = 1)
        ##plt.xticks(np.arange(0,max(profile),10), rotation = 'vertical')
        #plt.xlim(224,235)
        plt.plot(profile)

        for group in lines_settings.keys():            
            plt.axvline(x =  lines_settings[str(group)]["lower"], color = 'r', linestyle = 'dashed', linewidth = 1)
            plt.axvline(x =  lines_settings[str(group)]["upper"], color = 'r', linestyle = 'dashed', linewidth = 1)
        params = {'axes.labelsize': 18,
          'axes.titlesize': 24,
          'xtick.labelsize':8,
          'ytick.labelsize':16,
          'legend.fontsize': 18}
        plt.rcParams.update(params)

        plt.savefig('resolucionespacial.png')
        plt.show()
        
        #me guardo el rango de cada grupo
        group_profiles = {}
        group_contrast = {}
        for group in lines_settings.keys():
            #print(group)
            if group == "0":
                group_profiles[group] = profile[  lines_settings[group]["lower"]:lines_settings[group]["upper"]   ]
                #print(group_profiles[group])
                group_max = max( group_profiles[group]) #para normalizar al 100 luego
                #print(group_max)
                group_min = np.mean(np.array(group_profiles[group]))
                #print(group_min)
                group_contrast[group] = (group_max - group_min)/(group_max + group_min)
                #print(group_contrast[group])
            else:
                group_profiles[group] = profile[  lines_settings[group]["lower"]:lines_settings[group]["upper"]   ]
                #print(group_profiles[group])
                group_max = max( group_profiles[group])
                group_min = min( group_profiles[group])
                group_contrast[group] = (group_max - group_min)/(group_max + group_min)
                #print(group_contrast[group])
                

        group_max = max( group_profiles["1"]) #para normalizar al 100 luego
        #print(group_max)
        group_min = np.mean(np.array(group_profiles["0"]))
        #print(group_min)
        group_contrast["0"] = (group_max - group_min)/(group_max + group_min)
        #print(group_contrast[group])
                
        
        normalized_values = {}
        mtf = {}
        #https://www.normankoren.com/Tutorials/MTF5.html#quantitative
        #MTF(f) =  ( pi/4 ) * (  100%*C(f)/C(0)  )
        for group in lines_settings.keys():
            normalized_values[group] = np.round(  (group_contrast[group] / group_contrast["0"]) * 100  , 2)
            mtf[group] = np.round( (np.pi/4)*( group_contrast[group] / group_contrast["0"]) * 100  , 2)
       
        print(normalized_values)
        fig, ax = plt.subplots()   
        plt.xlabel('lp/cm')
        plt.ylabel('Contraste (%)')
        plt.ylim(0,100*1.1)
        plt.grid(linewidth = 1)
        
        for group in group_contrast.keys():
            plt.scatter(normalized_values.keys(), normalized_values.values(), color = 'red', marker = 'x')

        plt.savefig('resolucionespacial_contraste.png')
        plt.show()
        
        #solo funciona bien con 50%
        mtf_thresholds = [50]
        mtf_vals[k] = compute_mtf(list(normalized_values.keys()), list(mtf.values()), mtf_thresholds)

        return list(normalized_values.values()), normalized_values, group_contrast
        
#######################################
#
#            Funciones para Contraste
#
#######################################
def Contraste(head_params, dcm_files):
   
    imagenes = dcm_files
    num_imagenes = len(imagenes)

    for k in range(num_imagenes):
        
        # 2. Leer cada imagen       
        imagen = pydicom.dcmread(imagenes[k], force = True)

        pixel_data = imagen.pixel_array
        pixel_data_copy = imagen.pixel_array   
        
        print(' INFO **Contraste()** LEYENDO IMAGEN: ', imagenes[k])
        
         
        Hospital = imagen.InstitutionName
        ParteDelCuerpo = imagen.BodyPartExamined
        Equipo = imagen.ManufacturerModelName
        TamanoCorte = imagen.SliceThickness
        KV = imagen.KVP
        mAs = imagen.Exposure
        Kernel = imagen.ConvolutionKernel
        TamPixel = imagen.PixelSpacing
        tamPx = TamPixel[0]  # tamaño de Pixel en la exis.
        tamPy = TamPixel[1]  # tamaño de Pixel en la y.
        NumPixelsRows = imagen.Rows
        NumPixelsCols = imagen.Columns
        Intercept = imagen.RescaleIntercept
        Slope = imagen.RescaleSlope
         
        

        AIR = -1000
        PMP = -200
        LDPE = -100
        POLY = -35
        ACRYLIC = 120
        DELRIN = 340
        TEFLON = 990
        WATER = 0
        #Coeficientes de atenuacion lineal en cm-1 tomados a 70 kEV en base al protocolo
        roi_settings = {
            "Water": {
                "value": WATER,
                "angle": 90,
                "mu"   : 0.179,
                "color": "green",
                "theo" : "-7:7",
                "ref"  : 0,
            },  
            "Air": {
                "value": AIR,
                "angle": -90,
                "mu"   : 0,
                "color": "red",
                "theo" : "-1046:-986",
                "ref"  : -1000,
            },
            "PMP": {
                "value": PMP,
                "angle": -60,
                "mu"   : 0.157,
                "color": "blue",
                "theo" : "-220:-172",
                "ref"  : -200,
            },
            "LDPE": {
                "value": LDPE,
                "angle": 0,
                "mu"   : 0.174,
                "color": "orange",
                "theo" : "-121:-87",
                "ref"  : -100,
            },
            "Poly": {
                "value": POLY,
                "angle": 60,
                "mu"   : 0.188,
                "color": "yellow",
                "theo" : "-65:-29",
                "ref"  : -35,
            },
            "Acrylic": {
                "value": ACRYLIC,
                "angle": 120,
                "mu"   : 0.215,
                "color": "pink",
                "theo" : "92:137",
                "ref"  : 120,
            },
            "Delrin": {
                "value": DELRIN,
                "angle": -180,
                "mu"   : 0.245,
                "color": "black",
                "theo" : "315:345",
                "ref"  : 340,
            },
            "Teflon": {
                "value": TEFLON,
                "angle": -120,
                "mu"   : 0.363,
                "color": "cyan",
                "theo" : "915:955",
                "ref"  : 990,
            },
        }
        
        #Torax CT1 da error a -180, por lo que sea ponerlo a 180 y funciona
        if Equipo == "SOMATOM X.cite" and head_params == False:
            roi_settings["Delrin"]["angle"] = 180
            
        #CABEZA HR40
        if head_params == "Head":
            sigma = 5
            low_threshold = 1
            high_threshold = 10
        #TORAX BR40
        else:
            sigma = 5
            low_threshold = 1
            high_threshold = 10
            
        #detector de bordes
        edges = feature.canny(pixel_data, sigma, low_threshold, high_threshold, use_quantiles = False)
        #print(edges)
        plt.imshow(edges, cmap = 'gray')

        ##enumero y me quedo con los objetos de tras un procesado a partir del filtro
        thresh = np.mean(edges)
        bw = edges > thresh
        bw = segmentation.clear_border(bw, buffer_size=int(max(bw.shape) / 50))
        labeled_arr, num_roi = measure.label(bw, return_num=True)
        objects = measure.regionprops(labeled_arr)
        #print(objects)
        circle_objects = [obj for obj in objects if obj['eccentricity']<0.7 and obj['axis_major_length']>12/tamPx and obj['axis_major_length']<110/tamPx]

        fig, ax = plt.subplots(figsize=(10, 6))
        circle = {}
        roi_value = {}
        std_value = {}
        angle = {}
        for i in range(len(circle_objects)):

            #for each object, print out its centroid and print the object itself
            mask = np.zeros(shape=circle_objects[i]['image'].shape)
            mask[circle_objects[i]['image']]=1
            #mostrar que objetos ha detectado
            '''
            plt.figure(figsize=(2,2))
            plt.imshow(mask, cmap=plt.cm.gray)
            plt.show()
            '''
            
            #Get the coordinates of the circles and plot a rectangle around
            minr, minc, maxr, maxc = circle_objects[i].bbox
            #Apply a red rectangle enclosing each object of interest
            rect = mpatches.Rectangle((minc, minr), maxc - minc, maxr - minr, 
                           fill=False, edgecolor='red', linewidth=2)
            ax.add_patch(rect)
            ###wait = input("Press Enter to continue.")
            #row-col: coords
            coords, radius = analyze(circle_objects[i])
            ##print (i, coords, radius)
            a = int(coords[0])
            b = int(coords[1])
            r = int(3.5/tamPx)

            circle[i] = plt.Circle((b,a), r, color = "red", fill = False)
            ax.add_patch(circle[i])
            img = np.zeros(shape=pixel_data.shape)
            rr, cc = draw.disk((a, b), r, shape=pixel_data.shape)
            img[rr, cc] = 1
            
            std = pixel_data_copy[rr,cc]
            
            label_image = measure.label(img)
            props = measure.regionprops_table(label_image, pixel_data,
                          properties=['image_intensity', 'intensity_mean'])
                        
            #In HU units    
            roi_value[i] = np.round(props['intensity_mean'][0]*Slope + Intercept,2)
            std_value[i] = np.round(np.std(std),2)
            plt.text(b,a, str(i), color="red", fontsize=12)
            
            
            # Get image dimensions
            height, width = pixel_data.shape[:2]	
        	# Compute the center coordinates
            aa = height/2	
            bb = width/2
            #print(aa,bb)
            #print(a,b)         
            angle[i] = np.rad2deg(np.arctan2(-a+aa,b-bb))


        plt.imshow(pixel_data, cmap = 'gray')    
        
        roi_dict = {}
        std_dict = {}
        #inicializo el agua a 99999 para que no me problemas el bucle (a veces se detecta el agua o no, según se haya rellenado...)
        roi_dict["Water"] = 99999
        for roi in roi_settings.keys():
            for i in range(len(roi_value)):
                if np.abs(-angle[i]+int(roi_settings[roi]['angle'])) < 5:
                    roi_dict[roi] = roi_value[i]
                    std_dict[roi] = std_value[i]

                else:
                    pass

        print(" INFO **Contraste()** Valores ROI = ", roi_dict)
        print(" INFO **Contraste()** Std. Values = ", std_dict)
        
        
        return roi_dict, std_dict, roi_settings

           
def BajoContraste(head_params,dcm_files):
   
    imagenes = dcm_files
    num_imagenes = len(imagenes)

    for k in range(num_imagenes):
        
        # 2. Leer cada imagen       
        imagen = pydicom.dcmread(imagenes[k], force = True)

        pixel_data = imagen.pixel_array
        pixel_data_copy = imagen.pixel_array   
        
        print(' INFO **Contraste()** LEYENDO IMAGEN: ', imagenes[k])
        
         
        Hospital = imagen.InstitutionName
       #ParteDelCuerpo = imagen.BodyPartExamined
        Equipo = imagen.ManufacturerModelName
        TamanoCorte = imagen.SliceThickness
        KV = imagen.KVP
        #mAs = imagen.Exposure
        #Kernel = imagen.ConvolutionKernel
        TamPixel = imagen.PixelSpacing
        tamPx = TamPixel[0]  # tamaño de Pixel en la exis.
        tamPy = TamPixel[1]  # tamaño de Pixel en la y.
        NumPixelsRows = imagen.Rows
        NumPixelsCols = imagen.Columns
        Intercept = imagen.RescaleIntercept
        Slope = imagen.RescaleSlope
        #pixel_data_copy = Slope*pixel_data_copy+Intercept
        #print(pixel_data_copy.max(), pixel_data.max())
        plt.imshow(pixel_data, cmap = 'gray')
        
        # Get image dimensions
        height, width = pixel_data.shape[:2]	
    	# Compute the center coordinates
        aa = height/2	
        bb = width/2        
        roi_center_settings = {
            "Center": {
                "x": bb,
                "y": aa,
            },
        }
        
            
        roi_settings = {
            "15mm": {
                "diameter"  : 15,
                "angle"  : -93,
            },
            "9mm": {
                "diameter"  : 9,
                "angle"  : -112,
            },
            "8mm": {
                "diameter"  : 8,
                "angle"  : -129,
            },
            "7mm": {
                "diameter"  : 7,
                "angle"  : -141,
            },
            "6mm": {
                "diameter"  : 6,
                "angle"  : -156 ,
            },
            "5mm": {
                "diameter"  : 5,
                "angle"  : -168,
            },
            "4mm": {
                "diameter"  : 4,
                "angle"  : -177,
            },
            "3mm": {
                "diameter"  : 3,
                "angle"  : -179,
            },
            "2mm": {
                "diameter"  : 2,
                "angle"  : -180,
            },
        }
      
        # Aplicar el operador Canny para detección de bordes. Scharr es un buen filtro, pero se deja sin detectar uno de los materiales!
        #CABEZA HR40
        if head_params == "Head":
            sigma = 3
            low_threshold = 5
            high_threshold = 10
        #TORAX BR40
        else:
            sigma = 4.6
            low_threshold = 4.
            high_threshold = 7
        
        print('OSC', Equipo)
        #osc: procesado contraste (tanto para Head como para Body)
        if Equipo == "Spectral CT" or Equipo == "SOMATOM Definition AS+" or Equipo == "Ingenuity CT":
            pixel_data_rescaled = exposure.adjust_gamma(pixel_data, gamma=2.75, gain=1000) #CT3, CT2, CT4. El CT1 ya es reconstruido Hr40
        else:
            pixel_data_rescaled = pixel_data
        #plt.imshow(pixel_data_rescaled, cmap = 'gray')
        #detector de bordes
        edges = feature.canny(pixel_data_rescaled, sigma, low_threshold, high_threshold, use_quantiles = False)
        #print(edges)
        #plt.imshow(edges, cmap = 'gray')
        
        ##enumero y me quedo con los objetos de tras un procesado a partir del filtro
        thresh = np.mean(edges) 
        bw = edges > thresh
        bw = segmentation.clear_border(bw, buffer_size=int(max(bw.shape) / 50))
        labeled_arr, num_roi = measure.label(bw, return_num=True)
        #print(num_roi)
        #plt.imshow(bw)
        #propiedades de los objetos etiquetados
        objects = measure.regionprops(labeled_arr)
        #print(objects)
        ##descarto aquellos circulos que no me interesan a partir de la eccentricidad y ejes
        circle_objects = [obj for obj in objects if obj['eccentricity']<0.9 and obj['axis_major_length']<110/tamPx and obj['area'] > 5 ]
        fig, ax = plt.subplots()
        #print(circle_objects)
        circle = {}
        roi_value = {}
        angle = {}
        coords = {}
        radius = {}
        distance = {}
        for i in range(len(circle_objects)):
            #print(circle_objects[i]['area'])
            #for each object, print out its centroid and print the object itself
            mask = np.zeros(shape=circle_objects[i]['image'].shape)
            mask[circle_objects[i]['image']]=1
            #mostrar que objetos ha detectado
            #plt.figure(figsize=(2,2))
            #plt.imshow(mask, cmap=plt.cm.gray)
            #plt.show()
                    
            
            #Get the coordinates of the circles and plot a rectangle around
            minr, minc, maxr, maxc = circle_objects[i].bbox
            #Apply a red rectangle enclosing each object of interest
            rect = mpatches.Rectangle((minc, minr), maxc - minc, maxr - minr, 
                           fill=False, edgecolor='red', linewidth=2)
            ax.add_patch(rect)
            #wait = input("Press Enter to continue.")          
            
            #row-col: coords
            coords[i], radius[i] = analyze(circle_objects[i])
            #print (i, coords, radius)
            a = int(coords[i][0])
            b = int(coords[i][1])
            r = int(2/tamPx)

            circle[i] = plt.Circle((b,a), r, color = 'red', fill = False)
            ax.add_patch(circle[i])
            img = np.zeros(shape=pixel_data.shape)
            rr, cc = draw.disk((a, b), r, shape=pixel_data.shape)
            img[rr, cc] = 1
  
            label_image = measure.label(img)
            props = measure.regionprops_table(label_image, pixel_data,
                          properties=['image_intensity', 'intensity_mean'])
                        
            #In HU units    
            roi_value[i] =props['intensity_mean'][0]*Slope + Intercept
            plt.text(b+10,a+10, str(i), color="blue", fontsize=12)
            
            # Get image dimensions
            height, width = pixel_data.shape[:2]	
        	# Compute the center coordinates
            aa = height/2	
            bb = width/2
        
            angle[i] = np.rad2deg(np.arctan2(-a+aa,b-bb))
            distance[i] = (np.sqrt((aa-a)**2+(bb-b)**2))*tamPx

            
           
        #print(roi_dict, distance)
        circle_center = plt.Circle((bb,aa), r, color = 'red', fill = False)
        ax.add_patch(circle_center)
        img = np.zeros(shape=pixel_data.shape)
        rr, cc = draw.disk((aa, bb), r, shape=pixel_data.shape)
        img[rr, cc] = 1
  
        label_image = measure.label(img)
        props = measure.regionprops_table(label_image, pixel_data,
                      properties=['image_intensity', 'intensity_mean'])
                    
        #In HU units    
        circle_center_value =props['intensity_mean'][0]*Slope + Intercept
        
        contrast_percentage = [
            np.round(100/1000 * (( roi_value[i] - circle_center_value) / circle_center_value),3)
            for i in roi_value
        ]
        
      
        for i in range(len(contrast_percentage)):
            print(f' INFO **Contraste()** {i} ROI Contraste Porcentaje ( % )  = ', contrast_percentage[i])
            
        supra_dict = {}
        for roi in roi_settings.keys():
            for i in range(len(roi_value)):
              
                #angulo teorico y medido menor que 5 grados, y situados en el radio del grupo supra (en pixeles < 50)
                if  np.abs(angle[i] - roi_settings[roi]["angle"]) < 5 and distance[i] < 55:    #2.5  osc             
                    supra_dict[roi] = np.round(100/1000 * (( roi_value[i] - circle_center_value) / circle_center_value),3)
                else:
                    pass
        #print(supra_dict)
        plt.imshow(pixel_data_rescaled, cmap = 'gray')
        plt.savefig('bajo_contraste.png')
        plt.show()
        
        return contrast_percentage, supra_dict, fig
    
    
def plot_linearity(roi_value, std_value, roi_settings):
    
    fig = Figure(figsize=(12, 8))
    
    ax = fig.add_subplot(111)
    ax.set_ylim(-1500, 1500)
    ax.set_xlim(0, 0.4)
    ax.set_xlabel('Atenuación ($\mu = cm^{-1}$)')
    ax.set_ylabel('UH')
    ax.set_title("Linealidad Contraste")

    t = np.linspace(0, 600, 50)
    mu = {}
    
    #si el agua no está, lo quito de la tabla y del plot
    if roi_value["Water"] == 99999:
        roi_settings.pop("Water")
        roi_value.pop("Water")
    
    for roi in roi_settings.keys():
        ax.scatter(roi_settings[roi]['mu'], roi_value[roi], label=roi, color=roi_settings[roi]["color"], marker='x')
        mu[roi] = roi_settings[roi]['mu']
    
    slope, intercept, r, p, se = stats.linregress(list(mu.values()), list(roi_value.values()))
    ax.plot(t, intercept + slope * t, 'red', label='Ajuste')
    ax.legend()
    
    results = ['Escala de Contraste = %.6f UH/$\mathrm{cm}^{-1}$' % (1/slope)]
    textstr = '\n'.join(results)
    ax.text(0.25, 0.55, textstr, fontsize='large')

    return fig

def plot_resolution(normalized_values, group_contrast):
    
    fig = Figure(figsize=(12, 8))
    
    ax = fig.add_subplot(111)
    ax.set_ylim(0,100*1.1)
    #ax.set_xlim(0, 0.4)
    ax.set_xlabel('lp/cm')
    ax.set_ylabel('Contraste (%)')
    ax.set_title('Resolucion espacial')    
    ax.grid(linewidth = 1)

    
    for group in group_contrast.keys():
        ax.scatter(normalized_values.keys(), normalized_values.values(), color = 'red', marker = 'x')
    return fig

def plot_uniformity():
   
    fig = Figure(figsize=(12, 8))
    
    ax = fig.add_subplot(111)
    ax.set_ylim(0,100*1.1)
    #ax.set_xlim(0, 0.4)
    ax.set_xlabel('lp/cm')
    ax.set_ylabel('Contraste (%)')
    ax.set_title('Resolucion espacial')    
    ax.grid(linewidth = 1)

    
    for group in group_contrast.keys():
        ax.scatter(normalized_values.keys(), normalized_values.values(), color = 'red', marker = 'x')
    return fig
    

    
def create_xlsx2(roi_value, roi_settings, dcm_files):
    
    imagenes = dcm_files
    num_imagenes = len(imagenes)
    imagen = {}
    wb = Workbook()
    ws = wb.active
    #ws.append(["Materiales", "\mu (1/cm)", "UH", "Diferencia", "Estado"])

    for k in range(num_imagenes):

        imagen[k] = pydicom.dcmread(imagenes[k], force = True)
        materials = list(roi_value.keys())
        materials.insert(0, "Materiales")
        mu_theo = ["Coef. Atenuacion (1/cm)"]
        hu = list(roi_value.values())
        hu.insert(0, "HU")
        hu_theo = ["HU teorico"]
        hu_ref = ["HU referencia"]
        hu_diff = ["Diferencia HU"]
        state = ["Estado"]
        hu_lower = []
        hu_upper = []
        fill_color = []
        for roi in roi_value.keys():
            mu_theo = mu_theo + [roi_settings[roi]["mu"]]
            hu_theo = hu_theo + [roi_settings[roi]["theo"]]
            hu_ref = hu_ref + [roi_settings[roi]["ref"]]
            hu_lower = hu_lower + [roi_settings[roi]["theo"].split(":")[0]]
            hu_upper = hu_upper + [roi_settings[roi]["theo"].split(":")[1]]
        
        hu_diff = hu_diff + list(map(lambda a, b: np.round(a-b,2), hu[1:], hu_ref[1:]))
        #hu_diff = hu_diff + [a-b for a,b in zip(hu_ref, hu_theo)]

        for i, val in enumerate(hu[1:]):
            result = "Correcto" if int(val) >= int(hu_lower[i]) and int(val) <= int(hu_upper[i]) else "Incorrecto"    
            # Set fill color based on the result
            if result == "Correcto":
                fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
            else:
                fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
            fill_color = fill_color + [fill]
            state = state + [result] 
        
        ws.append(materials)
        ws.append(mu_theo)
        ws.append(hu)
        ws.append(hu_theo)
        ws.append(hu_ref)
        ws.append(hu_diff)
        ws.append(state)
           
            
    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", 
                               showFirstColumn=False,
                               showLastColumn=False, 
                               showRowStripes=False, 
                               showColumnStripes=False)
    ft = Font(bold=True)
    align = Alignment(horizontal="center", vertical="center", wrapText=True)
    for row in ws["A1:H1"]:
        for cell in row:
            cell.font = ft
    for row in ws["A1:H10"]:
        for cell in row:
            cell.alignment = align
    for row in ws["B7:H7"]:
        for i, cell in enumerate(row):
            cell.fill = fill_color[i]
            
    tab = Table(displayName="Table1", ref="A1:H1")
    
    # set the width of the column 
    for column_cells in ws.columns:
        new_column_length = max(len(str(cell.value)) for cell in column_cells)
        new_column_letter = (get_column_letter(column_cells[0].column))
        if new_column_length > 0:
            ws.column_dimensions[new_column_letter].width = new_column_length*1.23
                      
    tab.tableStyleInfo = style
    ws.add_table(tab)
    
    wb.save("table_contraste.xlsx")
    
    return hu[1:]


###########################################
#
#           Funciones para Uniformidad  ###
#
###########################################


def analyze_uniformity(edges):
    
    #Encontrar los contornos presentes en la imagen (con imagen previamente ajustada)
    contours = measure.find_contours(edges)
    #print(measure.perimeter(edges, neighborhood=4)/(2*np.pi))
    label_image = measure.label(edges)
    props = measure.regionprops_table(label_image, edges,
                  properties=['area', 'area_filled', 'area_bbox', 'centroid',
                              'perimeter', 'bbox'])
    #print(props)
    #Props es un array que arroja informacion de todos los contornos conexos e inconexos que encuentra Canny.
    #Si encuentra más de uno, nos interesa pillar el segundo. Si no, pues cogemos el unico que encuentra.
    try:
        coords = [props['centroid-0'][1],props['centroid-1'][1]] #(row-col) #osc
        radius_row = (props['bbox-2'][1]-props['bbox-0'][1])/2
        radius_col = (props['bbox-3'][1]-props['bbox-1'][1])/2
    except:
        coords = [props['centroid-0'][0],props['centroid-1'][0]] #(row-col) #osc
        radius_row = (props['bbox-2'][0]-props['bbox-0'][0])/2
        radius_col = (props['bbox-3'][0]-props['bbox-1'][0])/2
    radius = (radius_row + radius_col)/2
    
    return coords, radius

def plot_img_and_hist(image, axes, bins=256):
    """Plot an image along with its histogram and cumulative histogram. Esta funcion era para ver el histograma de la imagen al elegir parametros. En desuso."""
    #image = img_as_float(image)
    ax_img, ax_hist = axes
    ax_cdf = ax_hist.twinx()

    # Display image
    ax_img.imshow(image, cmap=plt.cm.gray)
    ax_img.set_axis_off()

    # Display histogram
    ax_hist.hist(image.ravel(), bins=bins, histtype='step', color='black')
    ax_hist.ticklabel_format(axis='y', style='scientific', scilimits=(0, 0))
    ax_hist.set_xlabel('Pixel intensity')
    ax_hist.set_xlim(0, 1)
    ax_hist.set_yticks([])

    # Display cumulative distribution
    img_cdf, bins = exposure.cumulative_distribution(image, bins)
    ax_cdf.plot(bins, img_cdf, 'r')
    ax_cdf.set_yticks([])

    return ax_img, ax_hist, ax_cdf

def Uniformity(dcm_files):
   
    imagenes = dcm_files
    num_imagenes = len(imagenes)

    for k in range(num_imagenes):
        # 2. Leer cada imagen
        
        imagen = pydicom.dcmread(imagenes[k], force = True)

        pixel_data = imagen.pixel_array
        pixel_data_copy = imagen.pixel_array 
        
        Hospital = imagen.InstitutionName
        ParteDelCuerpo = imagen.BodyPartExamined
        Equipo = imagen.ManufacturerModelName
        TamanoCorte = imagen.SliceThickness
        KV = imagen.KVP
        mAs = imagen.Exposure
        Kernel = imagen.ConvolutionKernel
        TamPixel = imagen.PixelSpacing
        tamPx = TamPixel[0]  # tamaño de Pixel en la exis.
        tamPy = TamPixel[1]  # tamaño de Pixel en la y.
        NumPixelsRows = imagen.Rows
        NumPixelsCols = imagen.Columns
        Intercept = imagen.RescaleIntercept
        Slope = imagen.RescaleSlope

        
        #Ajustar contraste
        if Equipo == "SOMATOM X.cite" or Equipo == "SOMATOM Definition AS+":
            pixel_data_rescaled = exposure.adjust_gamma(pixel_data, gamma=0.735, gain=100) #CT3, CT1 SIEMENS
        elif Equipo == "Spectral CT":
            pixel_data_rescaled = exposure.adjust_gamma(pixel_data, gamma=0.75, gain=1000) #CT4 Philips
        elif Equipo == "Ingenuity CT":
            pixel_data_rescaled = exposure.adjust_gamma(pixel_data, gamma=1, gain=1000) #CT2 Philips


        plt.imshow(pixel_data_rescaled, cmap = 'gray')
        
        # Aplicar el operador Canny para detección de bordes
        edges = feature.canny(pixel_data_rescaled, sigma=10.0, low_threshold=1500, high_threshold=2000, use_quantiles = False)
        plt.imshow(edges, cmap = 'gray')
        
        #coordenadas del borde detectado
        coords, r = analyze_uniformity(edges)
        a = int(coords[0])
        b = int(coords[1])
        r = int(r)
        #print(a, b, r)
        
        
        #GRAFICO
        fig, ax = plt.subplots()
        plt.imshow(pixel_data, cmap = 'gray') #Show the film at the top
        plt.xlabel('')
        plt.ylabel('')
        #plt.scatter(a,b, c='black', marker='x', s=60) 
        ##circ = plt.Circle((a,b), r, color = 'blue', fill = False) #osc
        circ = plt.Circle((b,a), r, color = 'blue', fill = False) #osc
        ax.add_patch(circ)
        #plt.show()


        print(' INFO **Uniformity()** LEYENDO IMAGEN: ', imagenes[k])

        
        #Protocolo Español: radio de 1 cm
        radius = int(10/tamPx) #en pixeles
        radiusy = int(10/tamPy) #en pixeles
        
        #Ojo con la notacion: ROI's a 1 cm del borde
        roi_settings = {
            "Top": {
                "x": b,
                "y": a-r+2*radius,
            },
            "Right": {
                "x": b+r-2*radius,
                "y": a,
            },
            "Bottom": {
                "x": b,
                "y": a+r-2*radius,

            },
            "Left": {
                "x": b-r+2*radius,
                "y": a,

            },
            "Center": {
                "x": b,
                "y": a,
            },
        }
           
        roi_value = {}
        circle_plt = {}
        
        # Get image dimensions
        height, width = pixel_data.shape[:2]	
   	    # Compute the center coordinates
        aa = height/2	
        bb = width/2  
        delta_aa = {}
        delta_bb = {}
        #Calcular correcciones al centro del borde tomando en cuenta el centro de la imagen
        for roi in roi_settings.keys():
            # Calculate the coordinates of the circle ROI   
            roi_x = 0
            roi_y = 0
            roi_x = roi_settings[roi]['x'] 
            roi_y = roi_settings[roi]['y']
            delta_aa[roi] = roi_y - aa
            delta_bb[roi] = roi_x - bb
            #print(roi_x, aa, delta_aa[roi])
        
        #Calcular y dibujar los ROI considerando las correcciones anteriores
        for roi in roi_settings.keys():
           
            # Calculate the coordinates of the circle ROI        
            roi_x = roi_settings[roi]['x'] 
            roi_y = roi_settings[roi]['y'] 
                    
  
            circle_plt[roi] = plt.Circle((roi_y - delta_aa["Center"], roi_x - delta_bb["Center"]), radius, color = 'black', fill = False)
            ax.add_patch(circle_plt[roi])
            
        #plt.show()
        #plt.imshow(pixel_data_copy, cmap = 'gray')
        plt.savefig('uniformidad.png')
        #plt.imshow(pixel_data, cmap = 'gray')

        
        circle = {}
        std_value = {}           
        #fig, ax = plt.subplots(figsize=(10, 6))
        for roi in roi_settings.keys():
            # Calculate the coordinates of the circle ROI   
            roi_x = 0
            roi_y = 0
            roi_x = roi_settings[roi]['x'] 
            roi_y = roi_settings[roi]['y']
            
            circle[roi] = plt.Circle((roi_y - delta_aa["Center"],roi_x - delta_bb["Center"]), radius, color = 'red', fill = False)
            ax.add_patch(circle[roi])
            img = np.zeros(shape=pixel_data.shape)
            rr, cc = draw.disk((roi_y - delta_aa["Center"], roi_x + delta_bb["Center"]), radius, shape=pixel_data.shape) #osc
            img[rr, cc] = 1
            
            label_image = measure.label(img)
            props = measure.regionprops_table(label_image, pixel_data_copy,
                          properties=['intensity_mean'])
            
            std = pixel_data_copy[rr,cc]
            #print(std)
            #print(props['intensity_mean'])
            
            #In HU units     
            roi_value[roi] = np.round(props['intensity_mean'][0]*Slope + Intercept,2)
            std_value[roi] = np.round(np.std(std),2)
            print(f' DEBUG **Uniformity()** {roi} ROI Value: ', roi_value[roi])
            print(f' DEBUG **Uniformity()** {roi} ROI Std. Value: ', std_value[roi])
           

        
            
        uis = [
            100 * (1- np.abs((roi_value[roi] - roi_value['Center']) / (roi_value['Center'] + 1000)))
            for roi in roi_settings.keys()
        ]
       
        abs_uis = np.abs(uis)
        ui = uis[np.argmax(abs_uis)]
        print(' INFO **Uniformity()** Uniformity Index= ', ui)
        
        #Checkear que |ROI(x) - ROI(Center)| < 5 HU
        diffs = [
            np.round((roi_value[roi] - roi_value['Center']),2)
            for roi in roi_settings.keys()
        ]
        
        abs_diffs = np.abs(diffs)
        diff = abs_diffs[np.argmax(abs_diffs)]
        print(f' INFO **Uniformity()** Maximum Difference ROI Value wrt ROI Center: ', diff)
        if diff < 5:
            print(f' INFO **Uniformity()** IN TOLERANCE ')
        else:
            print(f' INFO **Uniformity()** OUT OF TOLERANCE ')
            
        return roi_value, std_value, diffs, fig
    

# Funciones para MTF
def MTF(head_params, mtf_thresholds, info_file, dcm_files):
    # 1. Abrir las imágenes dcm.
    #print(' INFO **MTF()** LEYENDO IMAGEN: ', (os.path.dirname(os.getcwd()) + '/EjercicioMTF/imagenes/verificacion/Fm_hdn_rx_ct3Ccman.6.001.dcm'))
    #image_dir = os.path.dirname(os.getcwd()) + '/EjercicioMTF//imagenes/verificacion/Fm_hdn_rx_ct3Ccman.6.001.dcm'

    #print(dcm_files)
    #imagenes = [image_dir]
    imagenes = dcm_files
    num_imagenes = len(imagenes)
    
    mtfp_vals = {}
    info = pydicom.dcmread(info_file[0])
    
    for k in range(num_imagenes):
        # 2. Leer cada imagen
        #try:
        imagen = pydicom.dcmread(imagenes[k])
        pixel_data = imagen.pixel_array
        pixel_data_copy = imagen.pixel_array
        print(' INFO **MTF()** LEYENDO IMAGEN: ', imagenes[k])
        #except AttributeError:
        #    continue
        
        Hospital = info.InstitutionName
        ParteDelCuerpo = info.BodyPartExamined
        Equipo = info.ManufacturerModelName
        TamanoCorte = info.SliceThickness
        KV = info.KVP
        mAs = info.Exposure
        Kernel = info.ConvolutionKernel
        TamPixel = info.PixelSpacing
        tamPx = TamPixel[0]  # tamaño de Pixel en la exis.
        tamPy = TamPixel[1]  # tamaño de Pixel en la y.
        NumPixelsRows = info.Rows
        NumPixelsCols = info.Columns
        try:
            Pitch = info.SpiralPitchFactor
            CTDI = info.CTDIvol
        except:
            print('Bad Pitch Value. Probably dataset not available. Adjust the value manually to continue!')
            Pitch = 99999
            CTDI = 99999
            pass
            
    	# Get image dimensions
        height, width = pixel_data.shape[:2]	
    	# Compute the center coordinates
        roi_x = width/2 
        roi_y = height/2
        radius = int(150/tamPx) #en pixeles
        img = np.zeros(shape=pixel_data.shape)
        circle = plt.Circle((roi_y,roi_x), radius, color = 'red', fill = False)
        #ax.add_patch(circle)
        rr, cc = draw.disk((roi_x, roi_y), radius, shape=pixel_data.shape)
        img[rr, cc] = 1
        px = pixel_data[rr,cc]
        
        label_image = measure.label(img)
        props = measure.regionprops_table(label_image, pixel_data,
                      properties=['intensity_max','intensity_min'])


    
        # 3. Encontrar el maximo de la imagen
        image_max = ndi.maximum_filter(pixel_data,NumPixelsRows, mode = 'constant')

        image_pos = ndi.maximum_position(pixel_data,NumPixelsRows)

        #pixel_data_rescaled = filters.butterworth(pixel_data)
        plt.imshow(pixel_data, cmap = 'gray')
        # 4. Coordinadas del punto maximo
        # cuidado, estan invertidas las coordenadas( x->1, y->0)
        coordinates = feature.peak_local_max(pixel_data)
        #print(coordinates)
        Xm = coordinates[0][1]
        Ym = coordinates[0][0]
        #Xm = 261
        #Ym = 308
        
        # 5. Trazar un perfil entorno al máximo en la dirección exis
        A=20 
        lx=(Xm-A/2, Xm+A/2)
        ly=(Ym, Ym)
        start = (Ym, 0) #Start of the profile line
        end = (Xm, NumPixelsRows) #End of the profile line
        profile = measure.profile_line(pixel_data, (Ym, Xm - A/2), (Ym, Xm + A/2)) #Take the profile line
        
        # 5.1. GRAFICO DE LA IMAGEN DE PIXELES Y DEL PERFIL 
        fig, ax = plt.subplots(2, 1, figsize=(10, 10)) #Create the figures
        ax[0].imshow(pixel_data, cmap = 'gray') #Show the film at the top
        ax[0].plot(lx, ly, 'r-', lw=2) #Plot a red line across the film
        #ax[1].set_ylim(8000, 8600) #Plot a red line across the film        )
        ax[1].plot(profile)
        ax[1].grid()


        # 6. Perfil a mm
        xinf = (Xm-A/2) * tamPx
        xsup = (Xm+A/2) * tamPx
        profile_mm = profile
        fx = profile_mm
        fx = profile_mm / max(profile_mm)  # Normalizar el perfil.
        
        # 7. GRAFICO DEL PERFIL EN MM
        plt.figure(figsize=(12, 8))
        x = np.linspace(-(xsup-xinf)/2, (xsup-xinf)/2, len(fx))
        xp = np.linspace(-(xsup-xinf)/2, (xsup-xinf)/2, 5*len(fx)) #100??? it depends!!!
        fxp = fit_splines(x,xp,fx)
        plt.plot(x,fx, '*-', label='Perfil')
        plt.plot(xp,fxp, label = 'Ajuste')
        plt.xlabel('mm')
        plt.ylabel('f(x)')
        plt.title('PERFIL')
        plt.legend()
        

        # 8. CALCULO DE LA MTF PERFIL
        OTF = fft(fx)  # fft gives the zero frequency term at 0 component. While the others are positive and negative interval.
        OTFn = OTF[1:] #select all except the zero frequency term (because of non sense)

        MTF = np.abs(OTFn)  # valor absoluto.
        MTF = MTF/max(MTF)  # Normalizar y trasladar el cero al centro
        MTF = fftshift(MTF)  # Normalizar y trasladar el cero al centro
        
        # 8. CALCULO DE LA MTF AJUSTE
        OTFp = fft(fxp)  # fft gives the zero frequency term at 0 component. While the others are positive and negative interval.
        OTFnp = OTFp[1:] #select all except the zero frequency term (because of non sense)

        MTFp = np.abs(OTFnp)  # valor absoluto.
        MTFp = MTFp/max(MTFp)  # Normalizar y trasladar el cero al centro
        MTFp = fftshift(MTFp)  # Normalizar y trasladar el cero al centro

        
        # 8.1. Pasamos a espacio de frecuencias
        N = len(MTF)
        Np = len(MTFp)
        DX = np.abs(x[3]-x[2])  # distancia entre muestra (mm)
        DXp = np.abs(xp[3]-xp[2])  # distancia entre muestra (mm)
        F = 1/DX  # frecuencia espacial (1/mm)
        Fp = 1/DXp  # frecuencia espacial (1/mm)
        a = np.linspace(-N/2, N/2, N)  # puntos del dominio de frecuencia.
        ap = np.linspace(-Np/2, Np/2, Np)  # puntos del dominio de frecuencia.
        #a = np.arange(-N/2, N/2)  # puntos del dominio de frecuencia.
        #ap = np.arange(-Np/2, Np/2)  # puntos del dominio de frecuencia.

        Df = F / N  # intervalo de frecuencia (mm^-1)
        Dfp = Fp / Np  # intervalo de frecuencia (mm^-1)
        f = a*Df*10  # frecuencia ciclos/cm
        fp = ap*Dfp*10  # frecuencia ciclos/cm
        

        
        # 8.2. GRAFICO DE LA MTF
        plt.figure(figsize=(12, 8))
        plt.xlim(0,max(f)+1)
        plt.ylim(0,1.1)
        plt.plot(f, MTF, '*', label='MTF')
        plt.plot(fp, MTFp,'-*', label='MTFp')
        plt.grid(linewidth = 1)
        plt.yticks(np.arange(0,1.1,0.1))
        plt.xticks(np.arange(0,max(f),0.5))
        plt.xlabel('Frecuencia Espacial (ciclos/cm)')
        plt.ylabel('MTF')
        
        #VALUE OF MTF AT 10% and 50%
        #mtf_thresolds = [0.2,0.5]
        #mtf_thresolds = [0.5]
        try:
            mtf_vals = compute_mtf(f, MTF, mtf_thresholds)
            print(' INFO **MTF()** Valores de la MTF Perfil al 10%, 50%: ', mtf_vals)
        except:
            mtf_vals = []
            #print(' INFO **MTF()** Valores de la MTF Perfil al 10%, 50%: ', mtf_vals)
            print('Something went wrong. Probably not enough points in range!')
            pass
        
        mtfp_vals[k] = {}
        mtfp_vals[k] = compute_mtf(fp, MTFp, mtf_thresholds)
        #print(' INFO **MTF()** Valores de la MTF Ajuste al 10%, 50%: ', mtfp_vals[k])
        
        # 9. RESUMEN GRÁFICO DEL ANALISIS DE LA MTF
        fig, axs = plt.subplots(2, 2)
        fig.set_figheight(15)
        fig.set_figwidth(15)
        params = {'axes.labelsize': 18,
          'axes.titlesize': 18,
          'xtick.labelsize':14,
          'ytick.labelsize':14,
          'legend.fontsize': 14}
        plt.rcParams.update(params)
        #plt.style.use('Solarize_Light2')
        #print(plt.style.available)
        
        #CONFIGURACION RESUMEN
        strings = [["Hospital: ", Hospital], ["Equipo: ", Equipo], ["Tamaño de corte: ", str(TamanoCorte)], ["kV: ", str(KV)], ["mAs: ", str(mAs)], ["Kernel: ", Kernel] ]
        config_string = '\n'.join(''.join(x) for x in strings)
        #axs[0, 0].text(0.05, 0.90, config_string, fontsize = 19, va = 'top')
        #axs[0, 0].set_title("CONFIGURACION", pad = 20)
        axs[0, 0].set_axis_off()
        axs[0 ,0].set_facecolor("#1CC4AF")
        #PERFIL Y AJUSTE
        axs[1, 0].set_title("PERFIL Y AJUSTE")
        axs[1, 0].set_xlabel('mm')
        axs[1, 0].set_ylabel('f(x)')
        axs[1, 0].plot(x, fx, '*-', label='Perfil')
        axs[1, 0].plot(xp, fxp, '-', label='Ajuste')
        axs[1, 0].legend()
        #IMAGEN METF
        #axs[0, 1].set_title("IMAGEN PARA CALCULAR LA MTF", pad = 20)
        axs[0, 1].imshow(imagen.pixel_array, cmap='gray')
        axs[0, 1].plot(Xm, Ym, 'ro', markersize=5)
        #PLOT MTF
        #axs[1, 1].set_title("PLOT DE LA MTF")
        axs[1, 1].set_xlabel('Frecuencia Espacial (ciclos/cm)')
        axs[1, 1].set_ylabel('MTF')
        axs[1, 1].set_xlim(0,max(f))
        axs[1, 1].set_ylim(0, 1.1)
        axs[1, 1].grid(linewidth=2)
        axs[1, 1].plot(fp, MTFp,'-*', label='MTF', color = 'orange')
        props = dict(boxstyle='round', facecolor='wheat', alpha=0.2)
        results = ['f (MTF = %i%%) = %.2f $\mathrm{cm}^{-1}$' %(10,mtfp_vals[k][mtf_thresholds[0]]), 'f (MTF = %i%%) = %.2f $\mathrm{cm}^{-1}$' %(50,mtfp_vals[k][mtf_thresholds[1]])]
        #results = ['f (MTF = %i%%) = %.2f $\mathrm{cm}^{-1}$' %(50,mtfp_vals[k][0.5])]
        textstr = '\n'.join(results)
        axs[1,1].text(0.30, 0.90, textstr, transform=axs[1,1].transAxes, fontsize=14,
        verticalalignment='top', bbox=props)
        plt.savefig('%s.png' %(imagenes[k]))
        plt.show()

        print(mtfp_vals[k])
    
        return mtfp_vals[k], fig
    
    #######################################
#
#            Funciones para botones
#
#######################################

def procesar_mtf():
    # 01-04-24 exportar en funcion de la funcion que se ha llamado
    global last_function_called, data_to_export
    last_function_called = procesar_mtf
    # 01-04-24 exportar en funcion de la funcion que se ha llamado
            
    path = ruta_texto.get()
    anatomia = seleccion_anatomia.get()
    head_params = anatomia == 'Head'
    
    
    os.chdir(path)
    dcm_files = []
    info_file = []
    
    for file_name in os.listdir(path):
        try:
            if file_name.endswith(".DCM"):
                dcm_files.append(os.path.join(path, file_name))
                print(' INFO **MAIN()** CARGANDO IMAGEN: ', file_name)

            elif file_name.endswith(".dcm"):
                if "info" in file_name:
                    info_file.append(os.path.join(path, file_name))
                    print(' INFO **MAIN()** CARGANDO ARCHIVO DE INFORMACION: ', file_name)
                else: 
                    dcm_files.append(os.path.join(path, file_name))
                    print(' INFO **MAIN()** CARGANDO IMAGEN: ', file_name)

        except:
            pass
    
    
    mtf_thresholds = [0.1,0.5]
    #aqui necesito pillar el archivo sin comprimir para que encuentre bien el punto brillante
    #pero no contiene toda la info
    #asi que la info la pillo del archivo comprimido
    mtf, fig = MTF(head_params, mtf_thresholds, info_file, dcm_files)
    ##create_xlsx(mtf, mtf_thresholds, dcm_files)
    
    '''
    contraste, dict_val, groups = Resolucion(head_params, dcm_files)
    fig = plot_resolution(dict_val, groups)
    create_xlsx(contraste, dcm_files)
    '''
    
    # Limpia el área de mensajes y el Treeview si ya contiene datos
    area_mensajes.delete('1.0', tk.END)
    #for i in treeview.get_children():
    #    treeview.delete(i)
    # Si ya existe un canvas, primero lo eliminas
    for widget in grafico_frame.winfo_children():
        widget.destroy()
        
    canvas = FigureCanvasTkAgg(fig, master=grafico_frame)
    canvas.draw()
    canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)
    canvas.get_tk_widget().config(width=600, height=400) #600,500 osc
    
    #6-2-2024
    # Define el Treeview en GUI
    columns = ("MTF (%)", "Frecuencia de MTF (1/cm)")
    style = ttk.Style()
    style.configure("Treeview", rowheight=15)  # Ajusta la altura de las filas
    treeview = ttk.Treeview(app, columns=columns, show='headings')
    for col in columns:
        treeview.heading(col, text=col)
        treeview.column(col, anchor="center")
            
    treeview.configure(height=2)
    treeview.grid(row=7, column=0, columnspan=2, sticky='nsew')
    #6-2-2024
    
    
    # 01-04-24 exportar a excel
    data_to_export = list(mtf.values())

    # 01-04-24 exportar a excel

    #6-2-2024
    update_treeview(treeview, mtf_thresholds, data_to_export)
    #6-2-2024
    
    # Imprimiremos los valores seleccionados
    area_mensajes.insert(tk.END, f"Procesando imágenes en {path} para {anatomia}\n")
 

def procesar_resolucion():
    # 01-04-24 exportar en funcion de la funcion que se ha llamado
    global last_function_called, data_to_export
    last_function_called = procesar_resolucion
    # 01-04-24 exportar en funcion de la funcion que se ha llamado
            
    path = ruta_texto.get()
    anatomia = seleccion_anatomia.get()
    head_params = anatomia == 'Head'
    
    
    os.chdir(path)
    dcm_files = []
    
    
    for file_name in os.listdir(path):
        try:
            if file_name.endswith(".DCM"):
                dcm_files.append(os.path.join(path, file_name))
                print(' INFO **MAIN()** CARGANDO IMAGEN: ', file_name)
            elif file_name.endswith(".dcm"):
                dcm_files.append(os.path.join(path, file_name))
                print(' INFO **MAIN()** CARGANDO IMAGEN: ', file_name)
        except:
            pass
    
    
    contraste, dict_val, groups = Resolucion(head_params, dcm_files)
    fig = plot_resolution(dict_val, groups)
    create_xlsx(contraste, dcm_files)
    
    # Limpia el área de mensajes y el Treeview si ya contiene datos
    area_mensajes.delete('1.0', tk.END)
    #for i in treeview.get_children():
    #    treeview.delete(i)
    # Si ya existe un canvas, primero lo eliminas
    for widget in grafico_frame.winfo_children():
        widget.destroy()
        
    canvas = FigureCanvasTkAgg(fig, master=grafico_frame)
    canvas.draw()
    canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)
    canvas.get_tk_widget().config(width=600, height=500)
    
    #6-2-2024
    # Define el Treeview en GUI
    columns = ("lp/cm", "Contraste (%)")
    treeview = ttk.Treeview(app, columns=columns, show='headings')
    for col in columns:
        treeview.heading(col, text=col)
        treeview.column(col, anchor="center")
    treeview.grid(row=7, column=0, columnspan=2, sticky='nsew')
    #6-2-2024
    
    # 01-04-24 exportar a excel
    data_to_export = contraste
    rois = []
    # 01-04-24 exportar a excel

    #6-2-2024
    update_treeview(treeview, rois, contraste)
    #6-2-2024

    # Imprimiremos los valores seleccionados
    area_mensajes.insert(tk.END, f"Procesando imágenes en {path} para {anatomia}\n")


def procesar_contraste():
    # 01-04-24 exportar en funcion de la funcion que se ha llamado
    global last_function_called, data_to_export
    last_function_called = procesar_contraste
    # 01-04-24
    
    
    path = ruta_texto.get()
    anatomia = seleccion_anatomia.get()
    head_params = anatomia == 'Head'
    
    
    os.chdir(path)
    dcm_files = []
    
    
    for file_name in os.listdir(path):
        try:
            if file_name.endswith(".DCM"):
                dcm_files.append(os.path.join(path, file_name))
                print(' INFO **MAIN()** CARGANDO IMAGEN: ', file_name)
            elif file_name.endswith(".dcm"):
                dcm_files.append(os.path.join(path, file_name))
                print(' INFO **MAIN()** CARGANDO IMAGEN: ', file_name)
        except:
                pass
        
    ##########################

    roi_dict, std_dict, roi_settings = Contraste(head_params, dcm_files)
    fig = plot_linearity(roi_dict, std_dict, roi_settings)
    hu = create_xlsx2(roi_dict, roi_settings, dcm_files)
    
    # Limpia el área de mensajes y el Treeview si ya contiene datos
    area_mensajes.delete('1.0', tk.END)
    #for i in treeview_hu.get_children():
    #    treeview_hu.delete(i)

    # Si ya existe un canvas, primero lo eliminas
    for widget in grafico_frame.winfo_children():
        widget.destroy()
    
    canvas = FigureCanvasTkAgg(fig, master=grafico_frame)
    canvas.draw()
    canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)
    canvas.get_tk_widget().config(width=600, height=500)
        
        
    #03-04-2024
    # Define el Treeview en GUI
    columns_hu = ("Materiales", "HU")
    treeview_hu = ttk.Treeview(app, columns=columns_hu, show='headings')
    for col in columns_hu:
        treeview_hu.heading(col, text=col)
        treeview_hu.column(col, anchor="center")
    treeview_hu.grid(row=7, column=0, columnspan=2, sticky='nsew')
    #03-04-2024
    
    # 01-04-24 exportar a excel
    data_to_export = hu
    if "Water" in roi_dict.keys():
        materials = ["Water", "Air", "PMP", "LDPE", "Poly", "Acrylic", "Delrin", "Teflon", "Fecha"]
    else:
        materials = ["Air", "PMP", "LDPE", "Poly", "Acrylic", "Delrin", "Teflon", "Fecha"]
    # 01-04-24 exportar a excel
    update_treeview(treeview_hu, materials, hu)
    # Imprimiremos los valores seleccionados
    area_mensajes.insert(tk.END, f"Procesando imágenes en {path} para {anatomia}\n")


def procesar_uniformidad():
    # 01-04-24 exportar en funcion de la funcion que se ha llamado
    global last_function_called, data_to_export
    last_function_called = procesar_uniformidad
    # 01-04-24
    
    
    path = ruta_texto.get()
    anatomia = seleccion_anatomia.get()
    head_params = anatomia == 'Head'
    
    
    os.chdir(path)
    dcm_files = []
    
    
    for file_name in os.listdir(path):
        try:
            if file_name.endswith(".DCM"):
                dcm_files.append(os.path.join(path, file_name))
                print(' INFO **MAIN()** CARGANDO IMAGEN: ', file_name)
            elif file_name.endswith(".dcm"):
                dcm_files.append(os.path.join(path, file_name))
                print(' INFO **MAIN()** CARGANDO IMAGEN: ', file_name)
        except:
                pass
        
    ##########################
    
    std_threshold = 0
    if head_params == 'Head': std_threshold = 5
    elif head_params == 'Torax': std_threshold = 20 
    
    roi_value, std_value, diffs, fig = Uniformity(dcm_files)
    #roi_value_ruido, std_value_ruido= Ruido(std_threshold, dcm_files)
    ##create_xlsx(roi_value, std_value, std_threshold, diffs, dcm_files)
    ##create_xlsx_ruido(std_value_ruido, std_threshold, dcm_files)

    # Por ahora, solo imprimiremos los valores seleccionados
    area_mensajes.insert(tk.END, f"Procesando imágenes en {path} para {anatomia}\n")
    
    # Limpia el área de mensajes y el Treeview si ya contiene datos
    area_mensajes.delete('1.0', tk.END)
    #for i in treeview_hu.get_children():
    #    treeview_hu.delete(i)

    # Si ya existe un canvas, primero lo eliminas
    for widget in grafico_frame.winfo_children():
        widget.destroy()
    
    
    canvas = FigureCanvasTkAgg(fig, master=grafico_frame)
    canvas.draw()
    canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)
    canvas.get_tk_widget().config(width=600, height=500)
    
        
    #03-04-2024
    
    # Define el Treeview en GUI
    columns_hu = ("ROI", "Desviacion centro (HU)")
    treeview_hu = ttk.Treeview(app, columns=columns_hu, show='headings')
    for col in columns_hu:
        treeview_hu.heading(col, text=col)
        treeview_hu.column(col, anchor="center")
    treeview_hu.grid(row=7, column=0, columnspan=2, sticky='nsew')
    #03-04-2024
    
    # 01-04-24 exportar a excel
    data_to_export = diffs
    print(data_to_export)
    rois = ["ROI Top", "ROI Right", "ROI Bottom", "ROI Left", "ROI Center", "Fecha"]
    # 01-04-24 exportar a excel
    update_treeview(treeview_hu,rois, data_to_export)
    # Imprimiremos los valores seleccionados
    area_mensajes.insert(tk.END, f"Procesando imágenes en {path} para {anatomia}\n")

def procesar_bajocontraste():
    # 01-04-24 exportar en funcion de la funcion que se ha llamado
    global last_function_called, data_to_export
    last_function_called = procesar_bajocontraste
    # 01-04-24
    
    
    path = ruta_texto.get()
    anatomia = seleccion_anatomia.get()
    head_params = anatomia == 'Head'
    
    
    os.chdir(path)
    dcm_files = []
    
    
    for file_name in os.listdir(path):
        try:
            if file_name.endswith(".DCM"):
                dcm_files.append(os.path.join(path, file_name))
                print(' INFO **MAIN()** CARGANDO IMAGEN: ', file_name)
            elif file_name.endswith(".dcm"):
                dcm_files.append(os.path.join(path, file_name))
                print(' INFO **MAIN()** CARGANDO IMAGEN: ', file_name)
        except:
                pass
        
    ##########################
    
    contraste, supra_dict, fig = BajoContraste(head_params,dcm_files)
    #roi_value_ruido, std_value_ruido= Ruido(std_threshold, dcm_files)
    ##create_xlsx(roi_value, std_value, std_threshold, diffs, dcm_files)
    ##create_xlsx_ruido(std_value_ruido, std_threshold, dcm_files)

    # Por ahora, solo imprimiremos los valores seleccionados
    area_mensajes.insert(tk.END, f"Procesando imágenes en {path} para {anatomia}\n")
    
    # Limpia el área de mensajes y el Treeview si ya contiene datos
    area_mensajes.delete('1.0', tk.END)
    #for i in treeview_hu.get_children():
    #    treeview_hu.delete(i)

    # Si ya existe un canvas, primero lo eliminas
    for widget in grafico_frame.winfo_children():
        widget.destroy()
    
    
    canvas = FigureCanvasTkAgg(fig, master=grafico_frame)
    canvas.draw()
    canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)
    canvas.get_tk_widget().config(width=600, height=500)
    
        
    #03-04-2024
    
    # Define el Treeview en GUI
    columns_hu = ("ROI", "Contraste (%)")
    treeview_hu = ttk.Treeview(app, columns=columns_hu, show='headings')
    for col in columns_hu:
        treeview_hu.heading(col, text=col)
        treeview_hu.column(col, anchor="center")
    treeview_hu.grid(row=7, column=0, columnspan=2, sticky='nsew')
    #03-04-2024
    
    # 01-04-24 exportar a excel
    data_to_export = list(supra_dict.values())
    rois = list(supra_dict.keys())
    print(data_to_export)
    print(supra_dict)
    # 01-04-24 exportar a excel
    update_treeview(treeview_hu, rois, data_to_export)
    # Imprimiremos los valores seleccionados
    area_mensajes.insert(tk.END, f"Procesando imágenes en {path} para {anatomia}\n")

def seleccionar_carpeta():
    path = filedialog.askdirectory()
    ruta_texto.set(path)


def exportar_a_excel():

    path = ruta_texto.get()
    anatomia = seleccion_anatomia.get()
    head_params = anatomia == 'Head'

    # 01-04-24 exportar en funcion de la funcion que se ha llamado
    global last_function_called, data_to_export
    if last_function_called is None:
        print("No function has been called yet.")
        return

    # Prepare data based on the last function called
    if last_function_called == procesar_resolucion:
        print("Data from procesar_resolucion")
        if data_to_export[0] != "Contraste (%)":
            data_to_export.insert(0, "Contraste (%)")
        current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        data_to_export.append(current_date)
        lpcm = ["lp/cm", 0, 1 , 2, 3, 4, 5, 6, 7, 8, "Fecha"]
        print(data_to_export)
        try:
            wb = load_workbook("historico_resolucion.xlsx")
        except FileNotFoundError:
            wb = Workbook()

        ws = wb.active       
        row = 1    
        for i, value in enumerate(lpcm, start=1):
            ws.cell(row=row, column=i, value=value)

        # Encontrar la primera fila vacía en la columna A
        row = 2
        while ws.cell(row=row, column=1).value is not None: 
            row += 1
        for i, value in enumerate(data_to_export, start=1):
            ws.cell(row=row, column=i, value=value)
        wb.save("historico_resolucion.xlsx")
        
    # Prepare data based on the last function called
    elif last_function_called == procesar_mtf:
        print("Data from procesar_mtf")
        if data_to_export[0] != "MTF (%)":
            data_to_export.insert(0, "Frecuencia")
        current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        data_to_export.append(current_date)
        mtf = ["MTF (%)", 10, 50]
        print(data_to_export)
        try:
            wb = load_workbook("historico_mtf.xlsx")
        except FileNotFoundError:
            wb = Workbook()

        ws = wb.active       
        row = 1    
        for i, value in enumerate(mtf, start=1):
            ws.cell(row=row, column=i, value=value)

        # Encontrar la primera fila vacía en la columna A
        row = 2
        while ws.cell(row=row, column=1).value is not None: 
            row += 1
        for i, value in enumerate(data_to_export, start=1):
            ws.cell(row=row, column=i, value=value)
        wb.save("historico_mtf.xlsx")
        
    elif last_function_called == procesar_contraste:
        print("Data from procesar_contraste")
        
        if data_to_export[0] != "HU":
            data_to_export.insert(0, "HU")
        current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        data_to_export.append(current_date)
        try:
            materials = ["Materiales", "Water", "Air", "PMP", "LDPE", "Poly", "Acrylic", "Delrin", "Teflon", "Fecha"]
        except:
            materials = ["Materiales", "Air", "PMP", "LDPE", "Poly", "Acrylic", "Delrin", "Teflon", "Fecha"]
        print(data_to_export)
        try:
            wb = load_workbook("historico_contraste.xlsx")
        except FileNotFoundError:
            wb = Workbook()

        ws = wb.active       
        row = 1    
        for i, value in enumerate(materials, start=1):
            ws.cell(row=row, column=i, value=value)

        # Encontrar la primera fila vacía en la columna A
        row = 2
        while ws.cell(row=row, column=1).value is not None: 
            row += 1
        for i, value in enumerate(data_to_export, start=1):
            ws.cell(row=row, column=i, value=value)
        wb.save("historico_contraste.xlsx")
        
    elif last_function_called == procesar_uniformidad:
        print("Data from procesar_uniformidad")
        
        #if data_to_export[0] != "ROI":
        #    data_to_export.insert(0, "ROI")
        current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        data_to_export.append(current_date)
        rois = ["ROI Top", "ROI Right", "ROI Bottom", "ROI Left", "ROI Center", "Fecha"]
 
        print(data_to_export)
        try:
            wb = load_workbook("historico_uniformidad.xlsx")
        except FileNotFoundError:
            wb = Workbook()

        ws = wb.active       
        row = 1    
        for i, value in enumerate(rois, start=1):
            ws.cell(row=row, column=i, value=value)

        # Encontrar la primera fila vacía en la columna A
        row = 2
        while ws.cell(row=row, column=1).value is not None: 
            row += 1
        for i, value in enumerate(data_to_export, start=1):
            ws.cell(row=row, column=i, value=value)
        wb.save("historico_uniformidad.xlsx")
        
    elif last_function_called == procesar_bajocontraste:
        print("Data from procesar_bajocontraste")
        
        #if data_to_export[0] != "ROI":
        #    data_to_export.insert(0, "ROI")
        current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        rois = rois_bajocontraste
        rois.append("Fecha")
        
        data_to_export.append(current_date)
        
        print(data_to_export)
        try:
            wb = load_workbook("historico_bajocontraste.xlsx")
        except FileNotFoundError:
            wb = Workbook()

        ws = wb.active       
        row = 1    
        for i, value in enumerate(rois, start=1):
            ws.cell(row=row, column=i, value=value)

        # Encontrar la primera fila vacía en la columna A
        row = 2
        while ws.cell(row=row, column=1).value is not None: 
            row += 1
        for i, value in enumerate(data_to_export, start=1):
            ws.cell(row=row, column=i, value=value)
        wb.save("historico_bajocontraste.xlsx")
        

    else:
        print("Waiting for data")
    # 01-04-24 exportar en funcion de la funcion que se ha llamado

#####################################################################            
######################APP GUI #######################################
#####################################################################

global last_function_called

app = tk.Tk()
app.title("Procesador de Imágenes DICOM")


# Ruta de los archivos DICOM
tk.Label(app, text="Ruta de los Archivos DICOM:").grid(row=0, column=0, columnspan=2)
ruta_texto = tk.StringVar()
ruta_entrada = tk.Entry(app, textvariable=ruta_texto, width=100)
ruta_entrada.grid(row=1, column=0, columnspan=2)
tk.Button(app, text="Seleccionar Carpeta", command=seleccionar_carpeta).grid(row=2, column=0, columnspan=2)

# Selección de Anatomía
seleccion_anatomia = tk.StringVar(value='Head')
tk.Radiobutton(app, text="Head", variable=seleccion_anatomia, value='Head').grid(row=3, column=0)
tk.Radiobutton(app, text="Torax", variable=seleccion_anatomia, value='Torax').grid(row=3, column=1)

# Botón para procesar imágenes para MTF
tk.Button(app, text="Procesar MTF", command=procesar_mtf).grid(row=4, column=0)

# Botón para procesar imágenes para Resolución
tk.Button(app, text="Procesar Resolución", command=procesar_resolucion).grid(row=4, column=1)

# Botón para procesar imágenes para Uniformidad
tk.Button(app, text="Procesar Uniformidad", command=procesar_uniformidad).grid(row=5, column=0)

# Botón para procesar imágenes para Bajo Contraste
tk.Button(app, text="Procesar Bajo Contraste", command=procesar_bajocontraste).grid(row=5, column=1)

# Botón para procesar imágenes para Contraste
tk.Button(app, text="Procesar Contraste", command=procesar_contraste).grid(row=5, column=0, columnspan = 2)

#Botón para exportar a Excel histórico
tk.Button(app, text="Exportar", command=exportar_a_excel).grid(row=4, column=3)

# Área de Mensajes
area_mensajes = tk.Text(app, height=5, width=100)
area_mensajes.grid(row=6, column=0, columnspan=2)

# Contenedor para el gráfico
grafico_frame = tk.Frame(app)
grafico_frame.grid(row=8, column=0, columnspan=2)


app.mainloop()
        
