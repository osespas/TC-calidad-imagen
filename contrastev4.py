# -*- coding: utf-8 -*-
"""
Created on Thu Dec 21 18:15:06 2023

@author: oestpas
"""

# -*- coding: utf-8 -*-
"""
Created on Tue Nov 21 13:22:08 2023

@author: oestpas
"""

import os, sys, argparse, copy
import pydicom
from pydicom.data import get_testdata_files
import pylibjpeg
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import cv2
from skimage import measure, io, feature, draw, exposure, color, filters, segmentation
from skimage.filters.rank import maximum
from skimage.measure._regionprops import _RegionProperties
from scipy import ndimage as ndi
from scipy import interpolate, stats
from scipy.fft import fft, fftfreq, ifft, rfft, fftshift, fft2
from scipy.optimize import curve_fit,root
from scipy.interpolate import interp1d
import matplotlib.pylab as pylab
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles.colors import Color
import tkinter as tk
from tkinter import filedialog

def analyze(props):
    
    #Props es un array que arroja informacion de todos los contornos conexos e inconexos que encuentra Canny.
    #La componente-0 es la más grande (círculo interno tal y como está aplicado). PD. se podría hacer igualmente si pillamos el radio externo...
    coords = [props['centroid'][0],props['centroid'][1]] #(row-col)
    radius_row = (props['bbox'][2]-props['bbox'][0])/2
    radius_col = (props['bbox'][3]-props['bbox'][1])/2
    radius = (radius_row + radius_col)/2
    
    return coords, radius


def Contraste(head_params, dcm_files):
   
    imagenes = dcm_files
    num_imagenes = len(imagenes)

    for k in range(num_imagenes):
        
        # 2. Leer cada imagen       
        imagen = pydicom.dcmread(imagenes[k], force = True)

        pixel_data = imagen.pixel_array
        pixel_data_copy = imagen.pixel_array   
        
        print(' INFO **Contraste()** LEYENDO IMAGEN: ', imagenes[k])
        
         
        Hospital = imagen.InstitutionName
        ParteDelCuerpo = imagen.BodyPartExamined
        Equipo = imagen.ManufacturerModelName
        TamanoCorte = imagen.SliceThickness
        KV = imagen.KVP
        mAs = imagen.Exposure
        Kernel = imagen.ConvolutionKernel
        TamPixel = imagen.PixelSpacing
        tamPx = TamPixel[0]  # tamaño de Pixel en la exis.
        tamPy = TamPixel[1]  # tamaño de Pixel en la y.
        NumPixelsRows = imagen.Rows
        NumPixelsCols = imagen.Columns
        Intercept = imagen.RescaleIntercept
        Slope = imagen.RescaleSlope
        #pixel_data_copy = Slope*pixel_data_copy+Intercept
        #print(pixel_data_copy.max(), pixel_data.max())
         
        

        AIR = -1000
        PMP = -200
        LDPE = -100
        POLY = -35
        ACRYLIC = 120
        DELRIN = 340
        TEFLON = 990
        WATER = 0
        #Coeficientes de atenuacion lineal en cm-1 tomados a 70 kEV en base al protocolo
        roi_settings = {
            "Air": {
                "value": AIR,
                "angle": -90,
                "mu"   : 0,
                "color": "red",
                "theo" : "-1046:-986",
                "ref"  : -1000,
            },
            "PMP": {
                "value": PMP,
                "angle": -60,
                "mu"   : 0.157,
                "color": "blue",
                "theo" : "-220:-172",
                "ref"  : -200,
            },
            "LDPE": {
                "value": LDPE,
                "angle": 0,
                "mu"   : 0.174,
                "color": "orange",
                "theo" : "-121:-87",
                "ref"  : -100,
            },
            "Poly": {
                "value": POLY,
                "angle": 60,
                "mu"   : 0.188,
                "color": "yellow",
                "theo" : "-65:-29",
                "ref"  : -35,
            },
            "Acrylic": {
                "value": ACRYLIC,
                "angle": 120,
                "mu"   : 0.215,
                "color": "pink",
                "theo" : "92:137",
                "ref"  : 120,
            },
            "Delrin": {
                "value": DELRIN,
                "angle": -180, #tb 180 depende de si da error
                "mu"   : 0.245,
                "color": "black",
                "theo" : "315:345",
                "ref"  : 340,
            },
            "Teflon": {
                "value": TEFLON,
                "angle": -120,
                "mu"   : 0.363,
                "color": "cyan",
                "theo" : "915:955",
                "ref"  : 990,
            },
        }
        
        #CABEZA HR40
        if head_params == "Head":
            sigma = 5
            low_threshold = 1
            high_threshold = 10
        #TORAX BR40
        else:
            sigma = 5
            low_threshold = 1
            high_threshold = 10

        edges = feature.canny(pixel_data, sigma, low_threshold, high_threshold, use_quantiles = False)
        # Aplicar el operador Canny para detección de bordes. Scharr es un buen filtro, pero se deja sin detectar uno de los materiales!
        #edges = feature.canny(pixel_data, sigma=5.0, low_threshold=1, high_threshold=10, use_quantiles = False)

        plt.imshow(edges, cmap = 'gray')
        
        ##mean y otsu son filtros globales, los otros locales
        thresh = np.mean(edges)
        thresh_otsu = filters.threshold_otsu(edges)
        thresh_sauvola = filters.threshold_sauvola(edges)
        thresh_niblack = filters.threshold_niblack(edges, k=0.05)
        
        ##enumero y me quedo con los objetos de tras un procesado a partir del filtro
        bw = edges > thresh
        bw = segmentation.clear_border(bw, buffer_size=int(max(bw.shape) / 50))
        labeled_arr, num_roi = measure.label(bw, return_num=True)
        #labeled_arr_scharr, num_roi_scharr = measure.label(edges_scharr, return_num=True)
        #propiedades de los objetos etiquetados
        #objects = measure.regionprops_table(labeled_arr, bw, properties=['eccentricity','area', 'area_filled', 'area_bbox', 'centroid',
                     #'perimeter', 'bbox'])
        objects = measure.regionprops(labeled_arr)
        #objects_scharr = measure.regionprops(labeled_arr_scharr)
        
        ##descarto aquellos circulos que no me interesan a partir de la eccentricidad y ejes
        circle_objects = [obj for obj in objects if obj['eccentricity']<0.7 and obj['axis_major_length']>12/tamPx and obj['axis_major_length']<110/tamPx]
        #circle_objects_scharr = [obj for obj in objects_scharr if obj['centroid'][0] < 200 and obj['centroid'][1] < 300]
        fig, ax = plt.subplots(figsize=(10, 6))
        circle = {}
        roi_value = {}
        std_value = {}
        angle = {}
        for i in range(len(circle_objects)):

            #for each object, print out its centroid and print the object itself
            mask = np.zeros(shape=circle_objects[i]['image'].shape)
            mask[circle_objects[i]['image']]=1
            #mostrar que objetos ha detectado
            '''
            plt.figure(figsize=(2,2))
            plt.imshow(mask, cmap=plt.cm.gray)
            plt.show()
            '''
            
            #Get the coordinates of the circles and plot a rectangle around
            minr, minc, maxr, maxc = circle_objects[i].bbox
            #Apply a red rectangle enclosing each object of interest
            rect = mpatches.Rectangle((minc, minr), maxc - minc, maxr - minr, 
                           fill=False, edgecolor='red', linewidth=2)
            ax.add_patch(rect)
            ###wait = input("Press Enter to continue.")
            #row-col: coords
            coords, radius = analyze(circle_objects[i])
            ##print (i, coords, radius)
            a = int(coords[0])
            b = int(coords[1])
            r = int(3.5/tamPx)

            circle[i] = plt.Circle((b,a), r, color = "red", fill = False)
            ax.add_patch(circle[i])
            img = np.zeros(shape=pixel_data.shape)
            rr, cc = draw.disk((a, b), r, shape=pixel_data.shape)
            img[rr, cc] = 1
            
            std = pixel_data_copy[rr,cc]
            
            label_image = measure.label(img)
            props = measure.regionprops_table(label_image, pixel_data,
                          properties=['image_intensity', 'intensity_mean'])
                        
            #In HU units    
            roi_value[i] = np.round(props['intensity_mean'][0]*Slope + Intercept,2)
            std_value[i] = np.round(np.std(std),2)
            plt.text(b,a, str(i), color="red", fontsize=12)
            
            
            # Get image dimensions
            height, width = pixel_data.shape[:2]	
        	# Compute the center coordinates
            aa = height/2	
            bb = width/2
            #print(aa,bb)
            #print(a,b)         
            angle[i] = np.rad2deg(np.arctan2(-a+aa,b-bb))


        plt.imshow(pixel_data, cmap = 'gray')    
        
        roi_dict = {}
        std_dict = {}
        for roi in roi_settings.keys():
            for i in range(len(roi_value)):
                if np.abs(-angle[i]+int(roi_settings[roi]['angle'])) < 1:
                    roi_dict[roi] = roi_value[i]
                    std_dict[roi] = std_value[i]

                else:
                    pass

        print(" INFO **Contraste()** Valores ROI = ", roi_dict)
        print(" INFO **Contraste()** Std. Values = ", std_dict)
        
        
        return roi_dict, std_dict, roi_settings

           

def plot_linearity(roi_value, std_value, roi_settings):
    
    plt.figure(figsize=(12, 8))
    plt.ylim(-1500, 1500)
    plt.xlim(0,0.4)
    plt.xlabel('Atenuacion ($\mu = cm^{-1}$)')
    plt.ylabel('UH')    
    plt.title("Linealidad Contraste")
    t = np.linspace(0,600,50)
            
    mu = {}
    for roi in roi_settings.keys(): 
        plt.scatter(roi_settings[roi]['mu'],roi_value[roi], label = roi , color = roi_settings[roi]["color"], marker = 'x')
        mu[roi] = roi_settings[roi]['mu']
        #plt.errorbar(roi_settings[roi]['mu'], roi_value[roi], yerr= std_dict[roi] )


    slope, intercept, r, p, se = stats.linregress(list(mu.values()), list(roi_value.values()))
    print( f"R-squared: {r**2:.6f}")
    plt.plot(t, intercept + slope*t, 'red', label='Ajuste')
    plt.legend() 
    #textstr = f'{r**2}'
    #fig.text(0.30, 0.90, textstr, transform=axs[1,1].transAxes, fontsize=18,
    #verticalalignment='top', bbox=props)ç props = dict(boxstyle='round', facecolor='wheat', alpha=0.2)
    results = ['Escala de Contraste = %.6f UH/$\mathrm{cm}^{-1}$' %(1/slope)]
    #results = ['f (MTF = %i%%) = %.2f $\mathrm{cm}^{-1}$' %(50,mtfp_vals[k][0.5])]
    textstr = '\n'.join(results)
    plt.text(0.25, 0.55, textstr, fontsize = 'large')
    plt.savefig('linealidad_contraste.png')
    plt.show()
    
    
def create_xlsx(roi_value, roi_settings, dcm_files):
    
    imagenes = dcm_files
    num_imagenes = len(imagenes)
    imagen = {}
    wb = Workbook()
    ws = wb.active
    #ws.append(["Materiales", "\mu (1/cm)", "UH", "Diferencia", "Estado"])

    for k in range(num_imagenes):

        imagen[k] = pydicom.dcmread(imagenes[k], force = True)
        materials = list(roi_value.keys())
        materials.insert(0, "Materiales")
        mu_theo = ["Coef. Atenuacion (1/cm)"]
        hu = list(roi_value.values())
        hu.insert(0, "HU")
        hu_theo = ["HU teorico"]
        hu_ref = ["HU referencia"]
        hu_diff = ["Diferencia HU"]
        state = ["Estado"]
        hu_lower = []
        hu_upper = []
        fill_color = []
        for roi in roi_value.keys():
            mu_theo = mu_theo + [roi_settings[roi]["mu"]]
            hu_theo = hu_theo + [roi_settings[roi]["theo"]]
            hu_ref = hu_ref + [roi_settings[roi]["ref"]]
            hu_lower = hu_lower + [roi_settings[roi]["theo"].split(":")[0]]
            hu_upper = hu_upper + [roi_settings[roi]["theo"].split(":")[1]]
        
        hu_diff = hu_diff + list(map(lambda a, b: np.round(a-b,2), hu[1:], hu_ref[1:]))
        #hu_diff = hu_diff + [a-b for a,b in zip(hu_ref, hu_theo)]

        for i, val in enumerate(hu[1:]):
            result = "Correcto" if val >= int(hu_lower[i]) and val <= int(hu_upper[i]) else "Incorrecto"    
            # Set fill color based on the result
            if result == "Correcto":
                fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
            else:
                fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
            fill_color = fill_color + [fill]
            state = state + [result] 
        
        ws.append(materials)
        ws.append(mu_theo)
        ws.append(hu)
        ws.append(hu_theo)
        ws.append(hu_ref)
        ws.append(hu_diff)
        ws.append(state)
           
            
    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", 
                               showFirstColumn=False,
                               showLastColumn=False, 
                               showRowStripes=False, 
                               showColumnStripes=False)
    ft = Font(bold=True)
    align = Alignment(horizontal="center", vertical="center", wrapText=True)
    for row in ws["A1:H1"]:
        for cell in row:
            cell.font = ft
    for row in ws["A1:H10"]:
        for cell in row:
            cell.alignment = align
    for row in ws["B7:H7"]:
        for i, cell in enumerate(row):
            cell.fill = fill_color[i]
            
    tab = Table(displayName="Table1", ref="A1:H1")
    
    # set the width of the column 
    for column_cells in ws.columns:
        new_column_length = max(len(str(cell.value)) for cell in column_cells)
        new_column_letter = (get_column_letter(column_cells[0].column))
        if new_column_length > 0:
            ws.column_dimensions[new_column_letter].width = new_column_length*1.23
                      
    tab.tableStyleInfo = style
    ws.add_table(tab)
    
    wb.save("table_contraste.xlsx")
    

def procesar_imagenes():
    path = ruta_texto.get()
    anatomia = seleccion_anatomia.get()
    head_params = anatomia
    
    
    os.chdir(path)
    dcm_files = []
    
    
    for file_name in os.listdir(path):
        try:
            if file_name.endswith(".DCM"):
                dcm_files.append(os.path.join(path, file_name))
                print(' INFO **MAIN()** CARGANDO IMAGEN: ', file_name)
            elif file_name.endswith(".dcm"):
                dcm_files.append(os.path.join(path, file_name))
                print(' INFO **MAIN()** CARGANDO IMAGEN: ', file_name)
        except:
            pass
    
    
    roi_dict, std_dict, roi_settings = Contraste(head_params, dcm_files)
    plot_linearity(roi_dict, std_dict, roi_settings) 
    create_xlsx(roi_dict, roi_settings, dcm_files)

    # Por ahora, solo imprimiremos los valores seleccionados
    area_mensajes.insert(tk.END, f"Procesando imágenes en {path} para {anatomia}\n")

def seleccionar_carpeta():
    path = filedialog.askdirectory()
    ruta_texto.set(path)



app = tk.Tk()
app.title("Procesador de Imágenes DICOM")

# Ruta de los archivos DICOM
tk.Label(app, text="Ruta de los Archivos DICOM:").pack()
ruta_texto = tk.StringVar()
ruta_entrada = tk.Entry(app, textvariable=ruta_texto, width=50)
ruta_entrada.pack()
tk.Button(app, text="Seleccionar Carpeta", command=seleccionar_carpeta).pack()

# Selección de Anatomía
seleccion_anatomia = tk.StringVar(value='Head')
tk.Radiobutton(app, text="Head", variable=seleccion_anatomia, value='Head').pack()
tk.Radiobutton(app, text="Torax", variable=seleccion_anatomia, value='Torax').pack()

# Botón para procesar imágenes
tk.Button(app, text="Procesar Imágenes", command=procesar_imagenes).pack()

# Área de Mensajes
area_mensajes = tk.Text(app, height=10, width=50)
area_mensajes.pack()

app.mainloop()